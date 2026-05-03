#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
<<<<<<< HEAD
BERTA — Painel Operacional do Supervisor v3.6
Telas: Producao Diaria | Repetidos | Infancia | Calendario | Qualidade | Diario | Justificativas
Fonte de dados: Repositório (API)
Ajustes: compatibilidade com novas colunas do processador (PZERO, endereço completo)
=======
BERTA — Painel Operacional do Supervisor v7.9
Base: BASEBOT.csv + VIP (repetida/infância) do GitHub (SalaGpon/Berta-bot)
Telas: Diário | Produção | Repetidos (VIP) | Infância (VIP) | Calendário | Qualidade
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
"""

import re
<<<<<<< HEAD
import base64
import io
=======
import io
import time
import random
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
import requests
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
import streamlit as st
import traceback

# =============================================================================
<<<<<<< HEAD
=======
# LIMPAR CACHE AO INICIAR
# =============================================================================

if "cache_cleared" not in st.session_state:
    st.cache_data.clear()
    st.session_state.cache_cleared = True

# =============================================================================
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
# 1. CONFIGURAÇÃO DA PÁGINA
# =============================================================================

st.set_page_config(
    layout="wide",
    page_title="BERTA - Painel Operacional",
    page_icon="📡",
    initial_sidebar_state="expanded",
)

# =============================================================================
<<<<<<< HEAD
# 2. CSS GLOBAL
=======
# 2. CSS GLOBAL (mantido igual ao original)
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
# =============================================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');
<<<<<<< HEAD

html, body { margin: 0; padding: 0; }
* { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }

.stApp,
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="block-container"],
.main, .block-container,
section.main > div,
div[data-testid="stVerticalBlock"] {
    background-color: #f5f7fa !important;
    color: #1a2332 !important;
}

[data-testid="stSidebar"] {
    background-color: #ffffff !important;
    border-right: 1px solid #dde3ed !important;
=======
html, body { margin: 0; padding: 0; }
* { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }
.stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"], [data-testid="block-container"],
.main, .block-container, section.main > div, div[data-testid="stVerticalBlock"] {
    background-color: #f5f7fa !important; color: #1a2332 !important;
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
}
[data-testid="stSidebar"] { background-color: #ffffff !important; border-right: 1px solid #dde3ed !important; }
[data-testid="stSidebar"] * { color: #1a2332 !important; }
<<<<<<< HEAD
[data-testid="stSidebar"] hr { border-color: #dde3ed !important; }

[data-baseweb="select"] > div,
[data-baseweb="input"] > div {
    background-color: #ffffff !important;
    border-color: #dde3ed !important;
    color: #1a2332 !important;
}
[data-baseweb="select"] span,
[data-baseweb="select"] div { color: #1a2332 !important; }
[data-baseweb="popover"] { background: #ffffff !important; }
[role="listbox"] { background: #ffffff !important; }
[role="option"] { color: #1a2332 !important; }

[data-baseweb="tab-list"] {
    background-color: #ffffff !important;
    border-bottom: 2px solid #dde3ed !important;
}
[data-baseweb="tab"] {
    color: #64748b !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    padding: 10px 20px !important;
    background: transparent !important;
}
[aria-selected="true"] {
    color: #1e3a5f !important;
    border-bottom: 2px solid #1e3a5f !important;
    background: transparent !important;
}
[data-baseweb="tab-panel"] { background-color: transparent !important; }
[data-baseweb="tab-border"] { background-color: #dde3ed !important; }

.stDataFrame,
[data-testid="stDataFrameResizable"],
[data-testid="stDataFrameContainer"] {
    background-color: #ffffff !important;
    border: 1px solid #dde3ed !important;
    border-radius: 8px !important;
}

.stButton > button {
    background-color: #1e3a5f !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
}
.stButton > button:hover { background-color: #163050 !important; }

[data-baseweb="tag"] {
    background-color: #e8f0fa !important;
    color: #1e3a5f !important;
}

.stRadio label { color: #1a2332 !important; }

::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f5f7fa; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }

.kpi-card {
    background: #ffffff;
    border: 1px solid #dde3ed;
    border-radius: 10px;
    padding: 16px 14px;
    text-align: center;
    position: relative;
    overflow: hidden;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05);
    min-height: 100px;
}
.kpi-card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 3px;
    background: var(--accent, #1e3a5f);
}
.kpi-label {
    font-size: 10px;
    font-weight: 700;
    letter-spacing: 1.1px;
    text-transform: uppercase;
    color: #64748b;
    margin-bottom: 6px;
}
.kpi-value {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 26px;
    font-weight: 700;
    color: #1a2332;
    line-height: 1;
=======
[data-baseweb="select"] > div, [data-baseweb="input"] > div {
    background-color: #ffffff !important; border-color: #dde3ed !important;
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
}
.stButton > button { background-color: #1e3a5f !important; color: #ffffff !important; }
.kpi-card { background: #ffffff; border: 1px solid #dde3ed; border-radius: 10px; padding: 16px 14px;
    text-align: center; position: relative; box-shadow: 0 1px 4px rgba(0,0,0,0.05); min-height: 100px; }
.kpi-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: var(--accent, #1e3a5f); }
.kpi-label { font-size: 10px; font-weight: 700; text-transform: uppercase; color: #64748b; margin-bottom: 6px; }
.kpi-value { font-family: 'JetBrains Mono', monospace !important; font-size: 26px; font-weight: 700; color: #1a2332; }
.kpi-sub { font-size: 10px; color: #94a3b8; margin-top: 5px; }
.kpi-blue   { --accent: #1e3a5f; }
.kpi-green  { --accent: #16a34a; }
.kpi-yellow { --accent: #d97706; }
.kpi-red    { --accent: #dc2626; }
.kpi-purple { --accent: #7c3aed; }
<<<<<<< HEAD

.sec {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1.4px;
    text-transform: uppercase;
    color: #1e3a5f;
    border-left: 3px solid #1e3a5f;
    padding-left: 10px;
    margin: 22px 0 10px 0;
}

.ph {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 12px 0 18px 0;
    border-bottom: 1px solid #dde3ed;
    margin-bottom: 18px;
}
=======
.sec { font-size: 11px; font-weight: 700; text-transform: uppercase; color: #1e3a5f;
    border-left: 3px solid #1e3a5f; padding-left: 10px; margin: 22px 0 10px 0; }
.ph { display: flex; align-items: center; gap: 10px; padding: 12px 0 18px 0; border-bottom: 1px solid #dde3ed; margin-bottom: 18px; }
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
.ph h1 { font-size: 20px; font-weight: 700; color: #1a2332; margin: 0; }
.badge { background: #e8f0fa; border: 1px solid #bdd0ea; color: #1e3a5f; font-size: 10px;
    font-weight: 700; padding: 3px 10px; border-radius: 20px; text-transform: uppercase; }
.badge-sup { background: #f0f4f8; border-color: #cbd5e1; color: #475569; }
<<<<<<< HEAD

.banner-sup {
    background: #e8f0fa;
    border: 1px solid #bdd0ea;
    border-radius: 6px;
    padding: 8px 14px;
    margin-bottom: 16px;
    font-size: 12px;
    color: #1e3a5f;
}
=======
.qualidade-table { width: 100%; border-collapse: collapse; font-size: 13px; background: #ffffff;
    border-radius: 8px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.05); }
.qualidade-table th { background: #f8fafc; border-bottom: 2px solid #dde3ed; padding: 10px;
    text-align: left; font-weight: 600; color: #475569; }
.qualidade-table td { padding: 8px 10px; border-bottom: 1px solid #e2e8f0; }
.qualidade-table tr:hover td { background: #f8fafc; }
.info-box { background: #e8f0fa; border: 1px solid #bdd0ea; border-radius: 6px;
    padding: 10px 14px; margin-bottom: 12px; font-size: 12px; color: #1e3a5f; }
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
</style>
""", unsafe_allow_html=True)

# =============================================================================
# 3. CONSTANTES
# =============================================================================

<<<<<<< HEAD
_DIR = os.path.dirname(os.path.abspath(__file__))
CAMINHO_BASE_LOCAL = next(
    (p for p in [
        os.path.join(_DIR, "BASEBOT.csv"),
        os.path.join(_DIR, "bases", "BASEBOT.csv"),
    ] if os.path.exists(p)),
    None,
)

C = {
    "bg":          "#ffffff",
    "paper":       "#ffffff",
    "grid":        "#e8edf3",
    "txt":         "#475569",
    "navy":        "#1e3a5f",
    "navy2":       "#2d5a8e",
    "navy3":       "#4a7db5",
    "green":       "#16a34a",
    "yellow":      "#d97706",
    "red":         "#dc2626",
    "purple":      "#7c3aed",
}
=======
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
GITHUB_REPO = "SalaGpon/Berta-bot"
BASE_CSV = "BASEBOT.csv"
VIP_FOLDER = "bases/bucket/VIP"

>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
ROYAL = "#1e3a5f"
ROYAL_LIGHT = "#4a7db5"

_LYT = dict(
    paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
    font=dict(family="Inter, sans-serif", color="#475569", size=12),
    margin=dict(l=40, r=20, t=44, b=36),
    xaxis=dict(gridcolor="#e8edf3", linecolor="#e8edf3", zeroline=False),
    yaxis=dict(gridcolor="#e8edf3", linecolor="#e8edf3", zeroline=False),
    legend=dict(bgcolor="rgba(255,255,255,.9)", bordercolor="#e8edf3", borderwidth=1),
)

<<<<<<< HEAD
URL_JUSTIFICATIVA = "https://script.google.com/macros/s/AKfycbzBQjCcxSMrKBlUH3H6PY4I1AvF_XKTvJcHKXsHFT6HI7J-MONPeLXTEvZZOo5da4-y/exec"
URL_ATAS_DRIVE    = "https://script.google.com/macros/s/AKfycbyEw2E759E79Tpy3wbgZt8tip6F82aYtQiZ33-CoIpcExiUX5dUXynoBqYzJaa-fPk7DA/exec"

# =============================================================================
# 4. CARREGAMENTO DO DATAFRAME
# =============================================================================

@st.cache_data(ttl=300, show_spinner=False)
def carregar_base_repositorio():
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            st.warning("🔑 Token de acesso não configurado. Contate o administrador.")
            return None

        headers = {"Authorization": f"token {token}"}
        meta_url = "https://api.github.com/repos/SalaGpon/Berta-bot/contents/BASEBOT.csv"
        meta = requests.get(meta_url, headers=headers, timeout=30)
        if meta.status_code != 200:
            st.warning(f"Erro ao acessar repositório: HTTP {meta.status_code}")
            return None

        sha = meta.json().get("sha")
        blob_url = f"https://api.github.com/repos/SalaGpon/Berta-bot/git/blobs/{sha}"
        blob_resp = requests.get(blob_url, headers=headers, timeout=60)
        if blob_resp.status_code != 200:
            st.warning(f"Erro ao baixar arquivo: HTTP {blob_resp.status_code}")
            return None

        b64_content = blob_resp.json().get("content")
        if not b64_content:
            st.warning("Arquivo vazio no repositório.")
            return None

        texto = base64.b64decode(b64_content).decode("utf-8-sig", errors="replace")

        for sep in ["\t", ";", ","]:
            try:
                df = pd.read_csv(
                    io.StringIO(texto),
                    sep=sep,
                    dtype=str,
                    engine="python",
                    on_bad_lines="warn",
                    encoding="utf-8-sig"
                )
                if df.shape[1] > 5:
                    st.success(f"✅ Base carregada do Repositório (API) – separador detectado: '{sep}'")
                    return df
            except Exception:
                continue

        try:
            df = pd.read_csv(
                io.StringIO(texto),
                sep=None,
                dtype=str,
                engine="python",
                on_bad_lines="warn"
            )
            if df.shape[1] > 5:
                st.success("✅ Base carregada do Repositório (API) – separador automático")
                return df
        except Exception as e:
            st.warning(f"Não foi possível ler o arquivo do repositório: {e}")

        return None

    except Exception as e:
        st.warning(f"Erro ao carregar do repositório: {e}")
        return None


@st.cache_data(ttl=300)
def carregar_base_local():
    if not CAMINHO_BASE_LOCAL:
        return None
    try:
        for sep in ["\t", ";", ","]:
            try:
                df = pd.read_csv(CAMINHO_BASE_LOCAL, sep=sep, dtype=str, engine="python", on_bad_lines="warn")
                if df.shape[1] > 5:
                    st.success(f"✅ Base carregada do arquivo local – separador '{sep}'")
                    return df
            except Exception:
                continue
        df = pd.read_csv(CAMINHO_BASE_LOCAL, sep=None, dtype=str, engine="python", on_bad_lines="warn")
        if df.shape[1] > 5:
            st.success("✅ Base carregada do arquivo local – separador automático")
            return df
    except Exception as e:
        st.error(f"Erro ao carregar arquivo local: {e}")
    return None


def carregar_base():
    df = carregar_base_repositorio()
    if df is None:
        df = carregar_base_local()
    if df is not None:
        return _processar_df(df)
=======
_ESTADO_MAP = {
    'CONCLUIDO COM SUCESSO':  'CONCLUÍDO COM SUCESSO',
    'CONCLUIDO SEM SUCESSO':  'CONCLUÍDO SEM SUCESSO',
    'NAO ATRIBUIDO':          'NÃO ATRIBUÍDO',
    'ATRIBUIDO':              'ATRIBUÍDO',
    'EM EXECUCAO':            'EM EXECUÇÃO',
    'CANCELADO':              'CANCELADO',
    'RECEBIDO':               'RECEBIDO',
    'EM DESLOCAMENTO':        'EM DESLOCAMENTO',
}

# =============================================================================
# 4. FUNÇÕES DE CARREGAMENTO (BASEBOT + VIP)
# =============================================================================

def carregar_base_github(file_path, sep=";", dtype=str):
    """Baixa um arquivo CSV do GitHub usando token."""
    if not GITHUB_TOKEN:
        return None
    headers = {"Authorization": f"token {GITHUB_TOKEN}"}
    cache_buster = f"{int(time.time())}_{random.randint(10000, 99999)}"
    url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/{file_path}?v={cache_buster}"
    try:
        resp = requests.get(url, headers=headers, timeout=60)
        if resp.status_code == 200:
            df = pd.read_csv(io.StringIO(resp.text), sep=sep, dtype=dtype, low_memory=False)
            df.columns = df.columns.str.strip()
            return df
    except Exception:
        pass
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
    return None

def carregar_basebot():
    """Carrega o BASEBOT.csv."""
    return carregar_base_github(BASE_CSV, sep=";")

<<<<<<< HEAD
def _processar_df(df):
    df.columns = df.columns.str.strip()

    col_fim = "DH_FIM_EXEC_REAL"
    col_ab  = "AB_BA"
    col_estado = "ESTADO"
    col_macro  = "MACRO"
    col_tecnico = "TECNICO_EXECUTOR"
    col_territorio = "TERRITORIO"
    col_sa = "SA"
    col_gpon = "GPON"
    col_descricao = "DESCRICAO"
    col_obs = "OBSERVACAO"
    col_cod_fechamento = "COD_FECHAMENTO"

    for c in [col_fim, col_ab, col_estado, col_macro, col_tecnico, col_territorio, col_sa, col_gpon]:
        if c not in df.columns:
            raise KeyError(f"Coluna '{c}' não encontrada. Colunas disponíveis: {list(df.columns)}")

    df["FIM_DT"] = pd.to_datetime(df[col_fim], dayfirst=True, errors="coerce")
    df["AB_DT"]  = pd.to_datetime(df[col_ab],  dayfirst=True, errors="coerce")

    df["Estado"]  = df[col_estado].str.strip().str.upper()
    df["Macro Atividade"] = df[col_macro].str.strip().str.upper()

    df["NOME_TEC"] = df[col_tecnico].apply(
        lambda v: str(v).split(" - ")[0].strip().title() if pd.notna(v) and " - " in str(v) else str(v).strip().title()
    )

    def _extrair_cod(nome):
        if pd.isna(nome) or not nome:
            return ""
        m = re.search(r"(TR|TT|TC)\d+", str(nome), re.I)
        return m.group(0).upper() if m else ""
    df["CODIGO_TECNICO_EXTRAIDO"] = df[col_tecnico].apply(_extrair_cod)

    df["DIA_FIM"] = df["FIM_DT"].dt.date
    df["MES_FIM"] = df["FIM_DT"].dt.to_period("M")
    df["SEM_FIM"] = df["FIM_DT"].dt.isocalendar().week.astype("Int64")
    df["ANO_FIM"] = df["FIM_DT"].dt.year.astype("Int64")
    df["DIA_AB"]  = df["AB_DT"].dt.date
    df["MES_AB"]  = df["AB_DT"].dt.to_period("M")
    df["SEM_AB"]  = df["AB_DT"].dt.isocalendar().week.astype("Int64")

    # *** CORREÇÃO 1: derivar flags de conclusão a partir do ESTADO ***
    df["FLAG_CONCLUIDO_SUCESSO"] = (
        df["Estado"].str.strip().str.upper() == "CONCLUÍDO COM SUCESSO"
    ).map({True: "SIM", False: "NAO"})
    df["FLAG_CONCLUIDO_SEM_SUCESSO"] = (
        df["Estado"].str.strip().str.upper() == "CONCLUÍDO SEM SUCESSO"
    ).map({True: "SIM", False: "NAO"})

    df["FLAG_REPARO_VALIDO"] = (
        (df["Macro Atividade"] == "REP-FTTH") &
        (df["Estado"] == "CONCLUÍDO COM SUCESSO")
    ).map({True: "SIM", False: "NAO"})
    df["FLAG_INSTALACAO_VALIDA"] = (
        (df["Macro Atividade"] == "INST-FTTH") &
        (df["Estado"] == "CONCLUÍDO COM SUCESSO")
    ).map({True: "SIM", False: "NAO"})

    # Mapeamento dos flags do processador
    for _f, _default in [
        ("FLAG_REPETIDO", "NAO"),
        ("FLAG_INFANCIA", "NAO"),
        ("FLAG_INF", "NAO"),
        ("TEC_ANTERIOR", ""),
        ("IF_DIAS", ""),
        ("ALARMADO", "NAO"),
        ("CDOE", ""),
    ]:
        if _f not in df.columns:
            df[_f] = _default
        else:
            df[_f] = df[_f].fillna(_default)

    # NOVO: Mapear PZERO_10H e PZERO_15H para os nomes esperados pelo painel
    if "PZERO_10H" in df.columns:
        df["FLAG_P0_10_DIA"] = df["PZERO_10H"].fillna("NAO")
    else:
        df["FLAG_P0_10_DIA"] = "NAO"

    if "PZERO_15H" in df.columns:
        df["FLAG_P0_15_DIA"] = df["PZERO_15H"].fillna("NAO")
    else:
        df["FLAG_P0_15_DIA"] = "NAO"

    # NOVO: Colunas de endereço expandidas
    for _col, _map in [
        ("TIPO_LOGRADOURO", "Tipo Logradouro"),
        ("LOGRADOURO", "Logradouro"),
        ("BAIRRO", "Bairro"),
        ("NUMERO", "Número"),
        ("COMPLEMENTO_TIPO", "Complemento Tipo"),
        ("COMPLEMENTO_DESC", "Complemento Desc"),
        ("CEP", "CEP"),
        ("CIDADE", "Cidade"),
    ]:
        if _col in df.columns:
            df[_map] = df[_col].fillna("")
        elif _map not in df.columns:
            df[_map] = ""

    df["FLAG_REPETIDO_30D"] = df["FLAG_REPETIDO"]
    df["FLAG_INFANCIA_30D"]  = df["FLAG_INFANCIA"]

    if "FLAG_REPETIDO_ABERTO" not in df.columns:
        df["FLAG_REPETIDO_ABERTO"] = "NAO"

    df["Território de serviço: Nome"] = df[col_territorio]
    df["Número SA"] = df[col_sa]
    df["FSLOI_GPONAccess"] = df[col_gpon]
    df["Técnico Atribuído"] = df[col_tecnico]

    if "Descrição" not in df.columns:
        df["Descrição"] = df.get(col_descricao, "")
    if "Observação" not in df.columns:
        df["Observação"] = df.get(col_obs, "")
    if "Código de encerramento" not in df.columns:
        df["Código de encerramento"] = df.get(col_cod_fechamento, "")

    return df
=======
def carregar_vip_repetida(mes_ref):
    """Carrega o arquivo VIP de repetidos para o mês informado (ex: '2026-04')."""
    ano_mes = mes_ref.replace("-", "_")
    # padrão do nome: 2026_04_ANL_FTTH_REPETIDA_30_DIAS000000000000.csv
    padrao = f"{ano_mes}_ANL_FTTH_REPETIDA_30_DIAS"
    # Tentamos listar os arquivos via API do GitHub (mas é mais simples tentar padrão com 000...)
    # Vamos tentar algumas variações (0, 1, 2 zeros...). Normalmente é 12 zeros.
    for zeros in ["000000000000", "00000000000", "0000000000"]:
        nome = f"{padrao}{zeros}.csv"
        caminho = f"{VIP_FOLDER}/{nome}"
        df = carregar_base_github(caminho, sep=",", dtype=str)
        if df is not None:
            return df
    return None
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)

def carregar_vip_infancia(mes_ref):
    """Carrega o arquivo VIP de infância para o mês informado."""
    ano_mes = mes_ref.replace("-", "_")
    padrao = f"{ano_mes}_ANL_FTTH_INFANCIA_30_DIAS"
    for zeros in ["000000000000", "00000000000", "0000000000"]:
        nome = f"{padrao}{zeros}.csv"
        caminho = f"{VIP_FOLDER}/{nome}"
        df = carregar_base_github(caminho, sep=",", dtype=str)
        if df is not None:
            return df
    return None

def normalizar_gpon(gpon):
    if pd.isna(gpon) or not gpon:
        return ""
    return str(gpon).strip().upper()

<<<<<<< HEAD
def extrair_codigo_tecnico(tecnico_str) -> str:
    try:
        m = re.search(r'(TR\d+|TT\d+|TC\d+)', str(tecnico_str).strip())
        return m.group(1).upper() if m else ""
    except:
        return ""


@st.cache_data(ttl=600)
def carregar_equipes(df=None):
    if df is None:
        df = carregar_base()
    equipes = {}
    if df is not None:
        for _, row in df.iterrows():
            sup = str(row.get("SUPERVISOR", "")).strip()
            if not sup or sup.upper() in ("", "NAN", "NONE"):
                continue
            sup = sup.title()
            cod = str(row.get("TR", "")).strip().upper()
            if not cod:
                cod = extrair_codigo_tecnico(row.get("TECNICO_EXECUTOR", ""))
            if not cod:
                continue
            equipes.setdefault(sup, [])
            if cod not in equipes[sup]:
                equipes[sup].append(cod)

    if not equipes:
        st.info("Equipes não encontradas no BASEBOT.csv, tentando planilha de presença...")
        try:
            import openpyxl
            url_presenca = "https://raw.githubusercontent.com/SalaGpon/Berta-bot/main/Presen%C3%A7a.xlsx"
            r = requests.get(url_presenca, timeout=30)
            if r.status_code == 200:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                if rows:
                    headers = [cell.value for cell in ws[1]]
                    tr_col = next((i for i, h in enumerate(headers) if h and "TR" in str(h).upper()), None)
                    sup_col = next((i for i, h in enumerate(headers) if h and "SUPERVISOR" in str(h).upper()), None)
                    if tr_col is not None and sup_col is not None:
                        for row in rows:
                            tr = str(row[tr_col]).strip().upper() if row[tr_col] else None
                            sup = str(row[sup_col]).strip().title() if row[sup_col] else None
                            if tr and sup:
                                equipes.setdefault(sup, [])
                                if tr not in equipes[sup]:
                                    equipes[sup].append(tr)
                wb.close()
                if equipes:
                    st.success("✅ Equipes carregadas da planilha de presença")
        except Exception as e:
            st.warning(f"Não foi possível carregar a planilha de presença: {e}")

    return equipes


# =============================================================================
# 4B. CARREGAR MAPEAMENTO DE CAUSAS MACRO
# =============================================================================

@st.cache_data(ttl=3600)
def carregar_causas_macro():
    try:
        url = "https://raw.githubusercontent.com/SalaGpon/Berta-bot/main/Causas.xlsx"
        r = requests.get(url, timeout=30)
        if r.status_code != 200:
            raise FileNotFoundError("Arquivo de causas não encontrado.")
        df = pd.read_excel(io.BytesIO(r.content), sheet_name="Causas", dtype=str)
        df.columns = df.columns.str.strip()
        mapa = {}
        for _, row in df.iterrows():
            cod = str(row.get("cod_fechamento", "")).strip()
            causa = str(row.get("Causa Macro", "")).strip()
            if cod and causa:
                mapa[cod] = causa
        return mapa
    except Exception as e:
        st.warning("Não foi possível carregar a planilha de causas macro. Usando descrições originais.")
        return {}
=======
@st.cache_data(ttl=300)
def carregar_todas_bases(mes_ref):
    """Carrega BASEBOT, VIP repetida e VIP infância para o mês."""
    df_base = carregar_basebot()
    df_vip_rep = carregar_vip_repetida(mes_ref)
    df_vip_inf = carregar_vip_infancia(mes_ref)
    if df_base is None:
        st.error("❌ Não foi possível carregar BASEBOT.csv")
        return None, None, None
    return df_base, df_vip_rep, df_vip_inf

def processar_basebot(df):
    """Normaliza o BASEBOT (datas, estados, códigos)."""
    df = df.copy()
    if 'DH_FIM_EXEC_REAL' in df.columns:
        df['FIM_DT'] = pd.to_datetime(df['DH_FIM_EXEC_REAL'], dayfirst=True, errors='coerce')
        df['DIA_FIM'] = df['FIM_DT'].dt.date
        df['MES_FIM'] = df['FIM_DT'].dt.to_period('M')
        df['ANO_FIM'] = df['FIM_DT'].dt.year.astype('Int64')
        df['SEM_FIM'] = df['FIM_DT'].dt.isocalendar().week.astype('Int64')
    else:
        df['FIM_DT'] = pd.NaT
        df['DIA_FIM'] = None

    if 'AB_BA' in df.columns:
        df['AB_DT'] = pd.to_datetime(df['AB_BA'], dayfirst=True, errors='coerce')
        df['DIA_AB'] = df['AB_DT'].dt.date
    else:
        df['AB_DT'] = pd.NaT

    if 'INICIO_AG' in df.columns:
        df['INICIO_AG_DT'] = pd.to_datetime(df['INICIO_AG'], dayfirst=True, errors='coerce')
    else:
        df['INICIO_AG_DT'] = pd.NaT

    if 'ESTADO' in df.columns:
        df['ESTADO_NORM'] = df['ESTADO'].astype(str).str.strip().str.upper()
        df['ESTADO_NORM'] = df['ESTADO_NORM'].replace(_ESTADO_MAP)
    else:
        df['ESTADO_NORM'] = ''

    if 'MACRO' in df.columns:
        df['MACRO_NORM'] = df['MACRO'].astype(str).str.strip().str.upper()
    else:
        df['MACRO_NORM'] = ''

    if 'TR' in df.columns:
        df['COD_TEC'] = df['TR'].astype(str).str.strip().str.upper()
    else:
        df['COD_TEC'] = ''

    if 'NOME_FUNC' in df.columns and df['NOME_FUNC'].notna().any():
        df['NOME_TEC'] = df['NOME_FUNC'].astype(str).str.strip().str.title()
    elif 'TECNICO_EXECUTOR' in df.columns:
        df['NOME_TEC'] = df['TECNICO_EXECUTOR'].apply(
            lambda x: str(x).split(' - ')[0].strip().title() if pd.notna(x) else ""
        )
    else:
        df['NOME_TEC'] = ""

    mes_col = next((c for c in ['MÊS', 'MES', 'Mês', 'Mes', 'mês', 'mes'] if c in df.columns), None)
    if mes_col:
        df['MES_REF'] = df[mes_col].astype(str).str.strip()
    elif not df['FIM_DT'].isna().all():
        df['MES_REF'] = df['FIM_DT'].dt.strftime('%Y-%m').fillna('')
    else:
        df['MES_REF'] = ''

    for flag in ['FLAG_REPETIDO', 'FLAG_INFANCIA', 'ALARMADO', 'PZERO_10H', 'PZERO_15H']:
        if flag not in df.columns:
            df[flag] = 'NAO'
        else:
            df[flag] = df[flag].fillna('NAO').replace({'': 'NAO'}).astype(str).str.upper()

    if 'TEC_ANTERIOR' not in df.columns:
        df['TEC_ANTERIOR'] = pd.NA
    if 'GPON' not in df.columns:
        df['GPON'] = pd.NA
    if 'TERRITORIO' not in df.columns:
        df['TERRITORIO'] = ''

    if not df['FIM_DT'].isna().all():
        ultima_data = df['FIM_DT'].max().strftime('%d/%m/%Y %H:%M')
    elif not df['AB_DT'].isna().all():
        ultima_data = df['AB_DT'].max().strftime('%d/%m/%Y %H:%M')
    else:
        ultima_data = 'N/A'

    return df, ultima_data
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)


# =============================================================================
# 5. HELPERS
# =============================================================================

def _endereco_completo(row):
    parts = []
    logr = row.get("LOGRADOURO", "")
    if logr and pd.notna(logr) and str(logr).strip():
        parts.append(str(logr).strip())
    num = row.get("NUMERO", "")
    if num and pd.notna(num) and str(num).strip() and str(num).strip() != "0":
        parts.append(str(num).strip())
    bairro = row.get("BAIRRO", "")
    if bairro and pd.notna(bairro) and str(bairro).strip():
        parts.append(str(bairro).strip())
    cidade = row.get("CIDADE", "")
    if cidade and pd.notna(cidade) and str(cidade).strip():
        parts.append(str(cidade).strip())
    return ", ".join(parts) if parts else "Endereço não disponível"

def _kpi(label, valor, sub="", cls="kpi-blue"):
    s = f'<div class="kpi-sub">{sub}</div>' if sub else ""
    return f'<div class="kpi-card {cls}"><div class="kpi-label">{label}</div><div class="kpi-value">{valor}</div>{s}</div>'

def _sec(txt):
    st.markdown(f'<div class="sec">{txt}</div>', unsafe_allow_html=True)

def _header(icone, titulo, f):
    mes = f.get("mes", "")
    sup = f.get("supervisor", "")
    sup_b = f'<span class="badge badge-sup">👑 {sup}</span>' if sup else ""
    st.markdown(f'<div class="ph"><h1>{icone} {titulo}</h1><span class="badge">{mes}</span>{sup_b}</div>', unsafe_allow_html=True)

def _lyt(titulo="", h=360):
    return {**_LYT, "height": h, "title": dict(text=titulo, font=dict(size=13, color="#1a2332"), x=0.01)}

def _bar_h(y, x, color, titulo="", h=340, labels=None):
    fig = go.Figure(go.Bar(x=x, y=y, orientation="h", marker_color=color, text=labels, textposition="inside", textfont_color="white"))
    fig.update_layout(**_lyt(titulo, h))
    fig.update_layout(yaxis_autorange="reversed")
    return fig

<<<<<<< HEAD

def _ev_dual(x, bars, line, bcolor, titulo, h=300, meta=None):
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_bar(x=x, y=bars, name="Qtd", marker_color=bcolor)
    fig.add_scatter(x=x, y=line, name="Taxa%", mode="lines+markers",
                    line=dict(color=ROYAL_LIGHT, width=2), marker_size=6, secondary_y=True)
    if meta is not None:
        fig.add_hline(y=meta, line_dash="dash", line_color=C["yellow"],
                      annotation_text=f"Meta {meta}%", secondary_y=True)
    fig.update_layout(**_lyt(titulo, h))
    fig.update_yaxes(showgrid=False, secondary_y=True)
    return fig


def _endereco_completo(row):
    """Formata endereço completo com complemento e CEP."""
    parts = []
    tipo = row.get("Tipo Logradouro", "")
    if tipo and pd.notna(tipo) and str(tipo).strip():
        parts.append(str(tipo).strip())
    logr = row.get("Logradouro", "")
    if logr and pd.notna(logr) and str(logr).strip():
        parts.append(str(logr).strip())
    num = row.get("Número", "")
    if num and pd.notna(num) and str(num).strip():
        parts.append(str(num).strip())
    comp_tipo = row.get("Complemento Tipo", "")
    comp_desc = row.get("Complemento Desc", "")
    if comp_tipo and pd.notna(comp_tipo) and str(comp_tipo).strip():
        complemento = str(comp_tipo).strip()
        if comp_desc and pd.notna(comp_desc) and str(comp_desc).strip():
            complemento += " " + str(comp_desc).strip()
        parts.append(complemento)
    bairro = row.get("Bairro", "")
    if bairro and pd.notna(bairro) and str(bairro).strip():
        parts.append(str(bairro).strip())
    cidade = row.get("Cidade", "")
    if cidade and pd.notna(cidade) and str(cidade).strip():
        parts.append(str(cidade).strip())
    else:
        terr = row.get("Território de serviço: Nome", "")
        if terr and pd.notna(terr) and str(terr).strip():
            terr_str = str(terr).strip()
            if "." in terr_str:
                parts.append(terr_str.split(".")[1].strip())
            else:
                parts.append(terr_str)
    cep = row.get("CEP", "")
    if cep and pd.notna(cep) and str(cep).strip():
        parts.append("CEP " + str(cep).strip())
    return ", ".join(parts) if parts else "Endereço não disponível"


# =============================================================================
# 6. SIDEBAR
=======
# =============================================================================
# 6. SIDEBAR / FILTROS (extrai meses do BASEBOT)
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
# =============================================================================

def sidebar(df_base, ultima_data):
    with st.sidebar:
        st.markdown('<div style="text-align:center;padding:18px 0 10px"><div style="font-size:22px;font-weight:800;color:#1e3a5f;letter-spacing:2px">📡 BERTA</div><div style="font-size:10px;color:#64748b;font-weight:600">PAINEL OPERACIONAL</div></div>', unsafe_allow_html=True)
        st.divider()
<<<<<<< HEAD

        tela = st.radio("Tela",
            ["📅 Diario", "📊 Producao Diaria", "🔁 Repetidos", "👶 Infancia",
             "📆 Calendario", "🏆 Qualidade", "📝 Justificativas"],
            label_visibility="collapsed", key="nav")
=======
        tela = st.radio("Tela", ["📅 Diario", "📊 Producao Diaria", "🔁 Repetidos", "👶 Infancia", "📆 Calendario", "🏆 Qualidade"], label_visibility="collapsed", key="nav")
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
        st.divider()
        st.markdown("**Filtros**")
<<<<<<< HEAD
        df_atual = df

        meses = sorted(df_atual["MES_FIM"].dropna().astype(str).unique(), reverse=True)
        mes_padrao = datetime.now().strftime("%Y-%m")
        mes_idx = meses.index(mes_padrao) if mes_padrao in meses else 0
        mes = st.selectbox("📅 Mes", meses, index=mes_idx, key="f_mes")

        eq = carregar_equipes(df_atual)
        sups = sorted(eq.keys())
        if sups:
            sup = st.selectbox("👑 Supervisor", ["— Todos —"] + sups, key="f_sup")
            tecs_sup = eq.get(sup, []) if sup != "— Todos —" else []
            if tecs_sup:
                st.caption(f"Equipe: {len(tecs_sup)} tecnico(s)")
        else:
            sup, tecs_sup = "— Todos —", []

        terrs = sorted(df_atual["Território de serviço: Nome"].dropna().unique())
        terr  = st.multiselect("📍 Territorio", terrs, key="f_terr")

        pool = ([t for t in tecs_sup if t in df_atual["CODIGO_TECNICO_EXTRAIDO"].values]
                if tecs_sup else sorted(df_atual["CODIGO_TECNICO_EXTRAIDO"].dropna().unique()))
        tec  = st.multiselect("👤 Tecnico", pool, key="f_tec")

=======
        todos_meses = sorted(df_base["MES_REF"].dropna().replace('', pd.NA).dropna().unique(), reverse=True)
        if not todos_meses:
            st.warning("Nenhum mês disponível na base.")
            return {}
        mes_atual = datetime.now().strftime("%Y-%m")
        mes_idx = todos_meses.index(mes_atual) if mes_atual in todos_meses else 0
        mes = st.selectbox("📅 Mes", todos_meses, index=mes_idx, key="f_mes")
        # Supervisores e equipes (apenas para filtro, não afeta VIPs)
        equipes = {}
        if "SUPERVISOR" in df_base.columns and "TR" in df_base.columns:
            for _, row in df_base.iterrows():
                sup = str(row.get("SUPERVISOR", "")).strip()
                if sup and sup.upper() not in ("", "NAN", "NONE"):
                    tr = str(row.get("TR", "")).strip().upper()
                    if tr and tr not in ("NAN", ""):
                        equipes.setdefault(sup, set()).add(tr)
        equipes = {k: list(v) for k, v in equipes.items()}
        sups_originais = sorted(equipes.keys())
        sup_selecionado = st.selectbox("👑 Supervisor", ["— Todos —"] + sups_originais, key="f_sup") if sups_originais else "— Todos —"
        tecs_sup = equipes.get(sup_selecionado, []) if sup_selecionado != "— Todos —" else []
        terrs = sorted(df_base["TERRITORIO"].dropna().replace('', pd.NA).dropna().unique())
        terr = st.multiselect("📍 Territorio", terrs, key="f_terr")
        pool = [t for t in tecs_sup if t in df_base["COD_TEC"].values] if tecs_sup else sorted(df_base["COD_TEC"].dropna().unique())
        tec = st.multiselect("👤 Tecnico", pool, key="f_tec")
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
        st.divider()
        st.markdown(f'<div class="info-box">📅 Último registro na base:<br><strong>{ultima_data}</strong></div>', unsafe_allow_html=True)
        if st.button("🔄 Recarregar base", use_container_width=True):
            st.cache_data.clear()
            st.session_state.clear()
            st.rerun()
    return {"tela": tela, "mes": mes, "supervisor": sup_selecionado if sup_selecionado != "— Todos —" else "", "tecs_sup": tecs_sup, "territorios": terr, "tecnicos": tec}

def _filtrar(df, f):
    r = df[df["MES_REF"] == f["mes"]].copy()
    if f["tecs_sup"]:
        r = r[r["COD_TEC"].isin(f["tecs_sup"])]
    if f["territorios"] and "TERRITORIO" in r.columns:
        r = r[r["TERRITORIO"].isin(f["territorios"])]
    if f["tecnicos"]:
        r = r[r["COD_TEC"].isin(f["tecnicos"])]
    return r

def _escopo(df, f):
    if f["tecs_sup"]:
        return df[df["COD_TEC"].isin(f["tecs_sup"])].copy()
    return df


# =============================================================================
<<<<<<< HEAD
# 7. TELAS
=======
# 7. TELAS (com base VIP para repetidos/infância)
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
# =============================================================================

def tela_diario(df, f):
    _header("📅", "Controle do Dia", f)
    df_mes = df[df["MES_REF"] == f["mes"]].copy()
    datas = sorted(df_mes["FIM_DT"].dropna().dt.date.unique(), reverse=True)
    if not datas:
        st.warning(f"Nenhuma atividade concluída no mês {f['mes']}.")
        return
    dia_sel = st.date_input("Data de referência", value=datas[0], min_value=datas[-1], max_value=datas[0], key="dia_ref")
    atividades_dia = df_mes[df_mes["FIM_DT"].dt.date == dia_sel].copy()
    concluidas = atividades_dia[atividades_dia["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO"]
    sem_sucesso = atividades_dia[atividades_dia["ESTADO_NORM"] == "CONCLUÍDO SEM SUCESSO"]
    canceladas = atividades_dia[atividades_dia["ESTADO_NORM"] == "CANCELADO"]
    abertas = df_mes[((df_mes["AB_DT"].dt.date == dia_sel) | (df_mes["INICIO_AG_DT"].dt.date == dia_sel)) & (df_mes["FIM_DT"].isna()) & (~df_mes["ESTADO_NORM"].isin(["CONCLUÍDO COM SUCESSO", "CONCLUÍDO SEM SUCESSO", "CANCELADO"]))].copy()
    total_concluidas = len(concluidas)
    total_pendencias = len(sem_sucesso)
    tecs_ativos = concluidas["COD_TEC"].nunique()
    efic = round(total_concluidas / (total_concluidas + total_pendencias) * 100, 1) if (total_concluidas + total_pendencias) > 0 else 0
    cols = st.columns(5)
    for col, (lb, vl, sb, cl) in zip(cols, [("Concluídos", f"{total_concluidas}", "c/ sucesso", "kpi-blue"), ("Eficácia", f"{efic}%", "suc/total", "kpi-green" if efic >= 85 else "kpi-yellow"), ("Sem Sucesso", f"{total_pendencias}", "pendências", "kpi-red" if total_pendencias > 0 else "kpi-green"), ("Cancelados", f"{len(canceladas)}", "cancelados", "kpi-yellow"), ("Técnicos", f"{tecs_ativos}", "com atividade", "kpi-blue")]):
        col.markdown(_kpi(lb, vl, sb, cl), unsafe_allow_html=True)
    st.write("")
    if not concluidas.empty:
        _sec("✅ ATIVIDADES CONCLUÍDAS COM SUCESSO")
        prod_tec = concluidas.groupby("COD_TEC").agg(Nome=("NOME_TEC", lambda x: x.iloc[0] if not x.empty else ""), Total=("SA", "count"), INST=("MACRO_NORM", lambda x: (x == "INST-FTTH").sum()), REP=("MACRO_NORM", lambda x: (x == "REP-FTTH").sum())).reset_index()
        prod_tec = prod_tec.sort_values("Total", ascending=False).reset_index(drop=True).rename(columns={"COD_TEC": "TR"})
        st.dataframe(prod_tec, use_container_width=True, hide_index=True)
        with st.expander("📋 Ver lista detalhada"):
            cols_detalhe = ["SA", "MACRO", "COD_TEC", "NOME_TEC", "LOGRADOURO", "NUMERO", "BAIRRO", "CIDADE", "DESCRICAO"]
            cols_exist = [c for c in cols_detalhe if c in concluidas.columns]
            if cols_exist:
                st.dataframe(concluidas[cols_exist], use_container_width=True, hide_index=True)
    if not abertas.empty:
        _sec("🔄 ATIVIDADES EM ABERTO (NÃO CONCLUÍDAS)")
        abertas = abertas.sort_values(["AB_DT", "INICIO_AG_DT"])
        cols_aberto = ["SA", "MACRO", "COD_TEC", "NOME_TEC", "ESTADO", "LOGRADOURO", "NUMERO", "BAIRRO", "CIDADE"]
        cols_aberto_exist = [c for c in cols_aberto if c in abertas.columns]
        if cols_aberto_exist:
            st.dataframe(abertas[cols_aberto_exist], use_container_width=True, hide_index=True)
    if not sem_sucesso.empty:
        _sec("⚠️ PENDÊNCIAS (CONCLUÍDOS SEM SUCESSO)")
        cols_ss = [c for c in ["SA", "MACRO", "COD_TEC", "COD_FECHAMENTO", "DESCRICAO"] if c in sem_sucesso.columns]
        st.dataframe(sem_sucesso[cols_ss], use_container_width=True, hide_index=True)
    if atividades_dia.empty and abertas.empty:
        st.info("📭 Nenhuma atividade registrada para esta data.")

<<<<<<< HEAD
    _sec("Producao por Dia")
    prod = suc.groupby("DIA_FIM").agg(
        Total=("Número SA", "count"),
        Inst =("Macro Atividade", lambda x: (x=="INST-FTTH").sum()),
        Rep  =("Macro Atividade", lambda x: (x=="REP-FTTH").sum()),
    ).reset_index()
    ss = dm[dm["FLAG_CONCLUIDO_SEM_SUCESSO"]=="SIM"].groupby("DIA_FIM").size().reset_index(name="SS")
    prod = prod.merge(ss, on="DIA_FIM", how="left")
    prod["SS"]    = prod["SS"].fillna(0).astype(int)
    prod["Efic%"] = (prod["Total"]/(prod["Total"]+prod["SS"])*100).round(1)
    prod["Dia"]   = pd.to_datetime(prod["DIA_FIM"]).dt.strftime("%d/%m")

    fig = go.Figure()
    fig.add_bar(x=prod["Dia"], y=prod["Inst"], name="INST", marker_color=ROYAL)
    fig.add_bar(x=prod["Dia"], y=prod["Rep"],  name="REP",  marker_color=ROYAL_LIGHT)
    fig.update_layout(barmode="stack", showlegend=True, **_lyt("Producao Diaria - INST + REP", 320))
    st.plotly_chart(fig, use_container_width=True)

    st.dataframe(
        prod[["Dia","Total","Inst","Rep","SS","Efic%"]].rename(
            columns={"Inst":"INST","Rep":"REP","SS":"Sem Suc.","Efic%":"Eficacia%"}),
        use_container_width=True, hide_index=True,
        column_config={"Eficacia%": st.column_config.ProgressColumn(
            "Eficacia%", format="%.1f%%", min_value=0, max_value=100)})

    _sec("Pareto de Tecnicos")
    pt = suc.groupby(["CODIGO_TECNICO_EXTRAIDO","NOME_TEC"]).size().reset_index(name="Prod")
    pt = pt.sort_values("Prod", ascending=False).reset_index(drop=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Top 5 — Maior Producao**")
        t5 = pt.head(5)
        st.plotly_chart(
            _bar_h(t5["NOME_TEC"], t5["Prod"], ROYAL,
                   labels=t5["CODIGO_TECNICO_EXTRAIDO"], h=260),
            use_container_width=True)
    with c2:
        st.markdown("**Top 5 — Menor Producao**")
        b5 = pt.tail(5).sort_values("Prod")
        st.plotly_chart(
            _bar_h(b5["NOME_TEC"], b5["Prod"], ROYAL_LIGHT,
                   labels=b5["CODIGO_TECNICO_EXTRAIDO"], h=260),
            use_container_width=True)

    _sec("Ranking Completo")
    pt["#"] = range(1, len(pt)+1)
    st.dataframe(
        pt[["#","NOME_TEC","CODIGO_TECNICO_EXTRAIDO","Prod"]].rename(
            columns={"NOME_TEC":"Nome","CODIGO_TECNICO_EXTRAIDO":"TR","Prod":"Producao"}),
        use_container_width=True, hide_index=True,
        column_config={"Producao": st.column_config.ProgressColumn(
            "Producao", format="%d", min_value=0, max_value=int(pt["Prod"].max()))})

    _sec("Evolucoes")
    tab1, tab2 = st.tabs(["📅 Semanal","📆 Mensal"])
    suc_sc = ds[ds["FLAG_CONCLUIDO_SUCESSO"]=="SIM"].copy()
    with tab1:
        suc_sc["AW"] = suc_sc["ANO_FIM"].astype(str)+"-S"+suc_sc["SEM_FIM"].astype(str).str.zfill(2)
        ev = suc_sc.groupby("AW").size().reset_index(name="T").sort_values("AW").tail(12)
        fig2 = go.Figure(go.Scatter(x=ev["AW"],y=ev["T"],mode="lines+markers",
                                    line=dict(color=ROYAL,width=2),marker_size=7, marker_color=ROYAL))
        fig2.update_layout(**_lyt("Producao Semanal - ultimas 12 semanas",300))
        st.plotly_chart(fig2, use_container_width=True)
    with tab2:
        ev2 = suc_sc.groupby("MES_FIM").size().reset_index(name="T")
        ev2["MES_FIM"] = ev2["MES_FIM"].astype(str)
        ev2 = ev2.sort_values("MES_FIM").tail(12)
        fig3 = go.Figure(go.Bar(x=ev2["MES_FIM"],y=ev2["T"],marker_color=ROYAL))
        fig3.update_layout(**_lyt("Producao Mensal",300))
        st.plotly_chart(fig3, use_container_width=True)


def tela_repetidos(dm, ds, f):
    _header("🔁", "Repetidos", f)

    tem_vip = ("TEC_ANTERIOR" in dm.columns) and ("FLAG_REPETIDO" in dm.columns)

    den_df = dm[
        (dm["Macro Atividade"] == "REP-FTTH") &
        (dm["FLAG_CONCLUIDO_SUCESSO"] == "SIM")
    ]

    if tem_vip:
        num_df = dm[
            (dm["FLAG_REPETIDO"] == "SIM") &
            (dm["TEC_ANTERIOR"].notna()) &
            (dm["TEC_ANTERIOR"].str.strip() != "")
        ]
        fonte_label = "🏛️ Fonte: VIP Oficial (TEC_ANTERIOR + FLAG_REPETIDO)"
    else:
        gpons_rep, _, _ = _calcular_repetidos_gpon(ds, f["mes"])
        num_df = pd.DataFrame()
        fonte_label = "⚙️ Fonte: Cálculo Interno (GPON)"

    def _n(v): return str(v).split(" - ")[0].strip().title() if pd.notna(v) else ""

    if tem_vip:
        den_tec = den_df.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
            Nome=("Técnico Atribuído", lambda x: _n(x.iloc[0]) if len(x) else ""),
            Total=("Número SA", "count")).reset_index()
        num_tec = num_df.groupby("TEC_ANTERIOR").size().reset_index(name="Repetidos")
        num_tec.rename(columns={"TEC_ANTERIOR": "CODIGO_TECNICO_EXTRAIDO"}, inplace=True)
        tb = den_tec.merge(num_tec, on="CODIGO_TECNICO_EXTRAIDO", how="left")
        tb["Repetidos"] = tb["Repetidos"].fillna(0).astype(int)
        tb["Taxa%"] = (tb["Repetidos"] / tb["Total"].replace(0, 1) * 100).round(2)
        tb = tb.sort_values("Taxa%", ascending=False).reset_index(drop=True)
        tb = tb.rename(columns={"CODIGO_TECNICO_EXTRAIDO": "TR"})
    else:
        den_tec = den_df.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
            Nome=("Técnico Atribuído", lambda x: _n(x.iloc[0]) if len(x) else ""),
            Total=("Número SA", "count")).reset_index()
        rep_por_tec = {}
        for gpon, info in gpons_rep.items():
            tr = info.get("pai_tr", "")
            if tr:
                rep_por_tec[tr] = rep_por_tec.get(tr, 0) + 1
        num_tec_df = pd.DataFrame(rep_por_tec.items(), columns=["CODIGO_TECNICO_EXTRAIDO", "Repetidos"]) if rep_por_tec else pd.DataFrame(columns=["CODIGO_TECNICO_EXTRAIDO", "Repetidos"])
        tb = den_tec.merge(num_tec_df, on="CODIGO_TECNICO_EXTRAIDO", how="left")
        tb["Repetidos"] = tb["Repetidos"].fillna(0).astype(int)
        tb["Taxa%"] = (tb["Repetidos"] / tb["Total"].replace(0,1) * 100).round(2)
        tb = tb.sort_values("Taxa%", ascending=False).reset_index(drop=True)
        tb = tb.rename(columns={"CODIGO_TECNICO_EXTRAIDO": "TR"})

    total_rep_tabela = int(tb["Total"].sum())
    repetidos_tabela = int(tb["Repetidos"].sum())
    taxa_tabela = round(repetidos_tabela / total_rep_tabela * 100, 2) if total_rep_tabela > 0 else 0

    rep_ab = ds[ds["FLAG_REPETIDO_ABERTO"] == "SIM"] if "FLAG_REPETIDO_ABERTO" in ds.columns else pd.DataFrame()
    if tem_vip:
        rep_alrm = dm[(dm["FLAG_REPETIDO"] == "SIM") & (dm["ALARMADO"] == "SIM")]
        rep_alrm_count = len(rep_alrm)
    else:
        rep_alrm_count = 0

    cols = st.columns(5)
    for col, (lb,vl,sb,cl) in zip(cols,[
        ("Total Reparos", f"{total_rep_tabela:,}",   "abertos no mes",  "kpi-blue"),
        ("Repetidos",     f"{repetidos_tabela:,}",   "reparos filhos",  "kpi-red" if taxa_tabela>9 else "kpi-yellow"),
        ("Taxa %",        f"{taxa_tabela}%",          "meta: <= 9%",     "kpi-red" if taxa_tabela>9 else "kpi-green"),
        ("Em Garantia",   f"{len(rep_ab):,}",         "abertos 30d",     "kpi-yellow"),
        ("Alarmados",     f"{rep_alrm_count}",        "GPON alarmado",   "kpi-red" if rep_alrm_count>0 else "kpi-green"),
    ]):
        col.markdown(_kpi(lb,vl,sb,cl), unsafe_allow_html=True)

    st.markdown(
        f'<div style="font-size:11px;color:#64748b;margin:6px 0 14px 2px;">{fonte_label}</div>',
        unsafe_allow_html=True)

    _sec("Indicador por Tecnico")
    tb["Status"] = tb["Taxa%"].apply(lambda t: "🔴" if t>12 else "🟡" if t>9 else "🟢" if t>0 else "⚪")
    st.dataframe(tb[["Status","Nome","TR","Repetidos","Total","Taxa%"]], use_container_width=True, hide_index=True,
                 column_config={"Taxa%": st.column_config.ProgressColumn(
                     "Taxa%", format="%.1f%%", min_value=0,
                     max_value=max(float(tb["Taxa%"].max()) if not tb.empty else 1, 1))})

    if tem_vip and not num_df.empty:
        # *** CORREÇÃO 2: removida a seção Histórico Resumo dos GPONs Repetidos ***

        st.divider()
        _sec("Repetidos por Dia do Mês")
        if "DIA_FIM" in num_df.columns:
            diario = num_df.groupby("DIA_FIM").size().reset_index(name="Repetidos")
            diario["Dia"] = pd.to_datetime(diario["DIA_FIM"]).dt.strftime("%d/%m")
            fig_dia = make_subplots(specs=[[{"secondary_y": False}]])
            fig_dia.add_bar(x=diario["Dia"], y=diario["Repetidos"], name="Repetidos", marker_color=ROYAL)
            fig_dia.add_scatter(x=diario["Dia"], y=diario["Repetidos"], mode="lines+markers",
                                line=dict(color=ROYAL_LIGHT, width=2), marker=dict(size=6, color=ROYAL_LIGHT),
                                name="Tendência")
            fig_dia.update_layout(**_lyt("", 300))
            fig_dia.update_yaxes(title_text="Quantidade")
            st.plotly_chart(fig_dia, use_container_width=True)
        else:
            st.info("Dados insuficientes para gráfico diário.")

        _sec("Pareto de Causas dos Repetidos")
        mapa_causas = carregar_causas_macro()
        if mapa_causas:
            num_df["CAUSA_MACRO"] = num_df["Código de encerramento"].astype(str).str.strip().map(mapa_causas)
            num_df["CAUSA_MACRO"] = num_df["CAUSA_MACRO"].fillna("OUTROS")
        else:
            num_df["CAUSA_MACRO"] = num_df["Descrição"].str[:50]

        causas = num_df.groupby("CAUSA_MACRO").size().reset_index(name="Qtd")
        causas = causas.sort_values("Qtd", ascending=False).head(15)
        causas["Perc"] = (causas["Qtd"] / causas["Qtd"].sum() * 100).round(1)
        causas["Acum"] = causas["Perc"].cumsum()
        fig_causas = make_subplots(specs=[[{"secondary_y": True}]])
        fig_causas.add_bar(x=causas["CAUSA_MACRO"], y=causas["Qtd"],
                           name="Ocorrências", marker_color=ROYAL)
        fig_causas.add_scatter(x=causas["CAUSA_MACRO"], y=causas["Acum"],
                               name="% Acumulado", mode="lines+markers",
                               line=dict(color=ROYAL_LIGHT, width=2), secondary_y=True)
        fig_causas.update_layout(**_lyt("Pareto de Causas", 350))
        fig_causas.update_yaxes(showgrid=False, secondary_y=True)
        st.plotly_chart(fig_causas, use_container_width=True)

        _sec("Pareto de Técnicos — Repetidos")
        pareto_tec = tb[["Nome","TR","Repetidos"]].sort_values("Repetidos", ascending=False)
        pareto_tec["Perc"] = (pareto_tec["Repetidos"] / pareto_tec["Repetidos"].sum() * 100).round(1)
        pareto_tec["Acum"] = pareto_tec["Perc"].cumsum()
        fig_tec = make_subplots(specs=[[{"secondary_y": True}]])
        fig_tec.add_bar(x=pareto_tec["Nome"], y=pareto_tec["Repetidos"],
                        name="Repetidos", marker_color=ROYAL)
        fig_tec.add_scatter(x=pareto_tec["Nome"], y=pareto_tec["Acum"],
                            name="% Acumulado", mode="lines+markers",
                            line=dict(color=ROYAL_LIGHT, width=2), secondary_y=True)
        fig_tec.update_layout(**_lyt("Pareto de Técnicos", 400))
        fig_tec.update_yaxes(showgrid=False, secondary_y=True)
        st.plotly_chart(fig_tec, use_container_width=True)


def _calcular_repetidos_gpon(ds, mes_str):
    try:
        per = pd.Period(mes_str, freq="M")
        ano, mes = per.year, per.month
    except Exception:
        return {}, 0, pd.DataFrame()

    primeiro_dia = pd.Timestamp(ano, mes, 1)
    if mes == 12:
        ultimo_dia = pd.Timestamp(ano+1, 1, 1) - pd.Timedelta(days=1)
    else:
        ultimo_dia = pd.Timestamp(ano, mes+1, 1) - pd.Timedelta(days=1)

    df = ds.copy()
    if "FIM_DT" not in df.columns:
        df["FIM_DT"] = pd.to_datetime(df["DH_FIM_EXEC_REAL"], dayfirst=True, errors="coerce")
    if "AB_DT" not in df.columns:
        df["AB_DT"] = pd.to_datetime(df["AB_BA"], dayfirst=True, errors="coerce")

    df["_GPON"] = df["FSLOI_GPONAccess"].astype(str).str.strip().str.upper()

    df_rep = df[
        (df["Macro Atividade"] == "REP-FTTH") &
        (df["FLAG_CONCLUIDO_SUCESSO"] == "SIM") &
        (df["FIM_DT"].notna()) &
        (df["_GPON"].notna()) &
        (~df["_GPON"].isin(["", "NAN"]))
    ].copy()

    gpons_repetidos = {}
    for gpon, grupo in df_rep.groupby("_GPON"):
        grupo = grupo.sort_values("FIM_DT").reset_index(drop=True)
        if len(grupo) < 2:
            continue
        for i in range(len(grupo) - 1):
            pai  = grupo.iloc[i]
            filho = grupo.iloc[i+1]
            delta = (filho["FIM_DT"] - pai["FIM_DT"]).days
            ab_filho = filho.get("AB_DT")
            if pd.notna(ab_filho) and (ab_filho < primeiro_dia or ab_filho > ultimo_dia):
                continue
            if delta <= 30:
                if gpon not in gpons_repetidos:
                    gpons_repetidos[gpon] = {
                        "pai_tr"   : pai.get("CODIGO_TECNICO_EXTRAIDO", ""),
                        "pai_nome" : pai.get("NOME_TEC", ""),
                        "pai_sa"   : pai.get("Número SA", ""),
                        "pai_fim"  : pai["FIM_DT"],
                        "filho_tr" : filho.get("CODIGO_TECNICO_EXTRAIDO", ""),
                        "filho_sa" : filho.get("Número SA", ""),
                        "delta"    : delta,
                    }
                break

    den_df = df[
        (df["Macro Atividade"] == "REP-FTTH") &
        (df["FLAG_CONCLUIDO_SUCESSO"] == "SIM") &
        (df["AB_DT"].notna()) &
        (df["AB_DT"] >= primeiro_dia) &
        (df["AB_DT"] <= ultimo_dia)
    ]
    return gpons_repetidos, len(den_df), den_df


def tela_infancia(dm, ds, f):
    _header("👶", "Infancia", f)

    inst = dm[
        (dm["Macro Atividade"] == "INST-FTTH") &
        (dm["FLAG_CONCLUIDO_SUCESSO"] == "SIM")
    ]

    tem_vip_inf = "FLAG_INFANCIA" in inst.columns

    if tem_vip_inf:
        inf = inst[inst["FLAG_INFANCIA"] == "SIM"]
        fonte_label = "🏛️ Fonte: VIP Oficial"
    else:
        rep = ds[(ds["Macro Atividade"]=="REP-FTTH") & (ds["FLAG_CONCLUIDO_SUCESSO"]=="SIM")]
        inf_list = []
        for _, irow in inst.iterrows():
            gpon = irow["FSLOI_GPONAccess"]
            if pd.isna(gpon): continue
            fim_inst = irow["FIM_DT"]
            if pd.isna(fim_inst): continue
            limite = fim_inst + pd.Timedelta(days=30)
            filhos = rep[(rep["FSLOI_GPONAccess"]==gpon) & (rep["FIM_DT"]>fim_inst) & (rep["FIM_DT"]<=limite)]
            if not filhos.empty:
                inf_list.append(irow)
        inf = pd.DataFrame(inf_list) if inf_list else pd.DataFrame()
        fonte_label = "⚙️ Fonte: Cálculo Interno"

    taxa = round(len(inf)/len(inst)*100,2) if len(inst)>0 else 0
    estados_ab = ["ATRIBUÍDO","NÃO ATRIBUÍDO","RECEBIDO","EM EXECUÇÃO","EM DESLOCAMENTO"]
    gpons = set(inf["FSLOI_GPONAccess"].dropna().str.upper())
    rep_ab = ds[(ds["Macro Atividade"]=="REP-FTTH") &
                (ds["Estado"].isin(estados_ab)) &
                (ds["FSLOI_GPONAccess"].str.upper().isin(gpons))] if gpons else pd.DataFrame()
    inf_alrm = inf[inf["ALARMADO"]=="SIM"] if "ALARMADO" in inf.columns else pd.DataFrame()

    cols = st.columns(5)
    for col,(lb,vl,sb,cl) in zip(cols,[
        ("Total Inst.",  f"{len(inst):,}", "INST concluidas",  "kpi-blue"),
        ("Infancia",     f"{len(inf):,}",  "reparo em 30d",    "kpi-red" if taxa>5 else "kpi-yellow"),
        ("Taxa %",       f"{taxa}%",       "meta: <= 5%",      "kpi-red" if taxa>5 else "kpi-green"),
        ("Inf. Aberta",  f"{len(rep_ab):,}","reparo aberto",   "kpi-yellow"),
        ("Alarmados",    f"{len(inf_alrm):,}","GPON alarmado", "kpi-red" if len(inf_alrm)>0 else "kpi-green"),
    ]):
        col.markdown(_kpi(lb,vl,sb,cl), unsafe_allow_html=True)

    st.markdown(
        f'<div style="font-size:11px;color:#64748b;margin:6px 0 14px 2px;">{fonte_label}</div>',
        unsafe_allow_html=True)

    _sec("Indicador por Tecnico")
    def _n(v): return str(v).split(" - ")[0].strip().title() if pd.notna(v) else ""

    di = inst.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
        Total=("Número SA","count"),
        Nome=("Técnico Atribuído", lambda x: _n(x.iloc[0]) if len(x) else "")
    ).reset_index()
    if tem_vip_inf:
        ni = inf.groupby("CODIGO_TECNICO_EXTRAIDO").size().reset_index(name="Infancia")
    else:
        ni = inf.groupby("CODIGO_TECNICO_EXTRAIDO").size().reset_index(name="Infancia") if not inf.empty else pd.DataFrame(columns=["CODIGO_TECNICO_EXTRAIDO","Infancia"])

    tb = di.merge(ni, on="CODIGO_TECNICO_EXTRAIDO", how="left")
    tb["Infancia"] = tb["Infancia"].fillna(0).astype(int)
    tb["Taxa%"] = (tb["Infancia"] / tb["Total"].replace(0,1) * 100).round(2)
    tb = tb.sort_values("Taxa%", ascending=False).reset_index(drop=True)
    tb = tb.rename(columns={"CODIGO_TECNICO_EXTRAIDO":"TR"})
    tb["Status"] = tb["Taxa%"].apply(lambda t: "🔴" if t>8 else "🟡" if t>5 else "🟢" if t>0 else "⚪")
    st.dataframe(tb[["Status","Nome","TR","Infancia","Total","Taxa%"]], use_container_width=True, hide_index=True,
                 column_config={"Taxa%": st.column_config.ProgressColumn(
                     "Taxa%", format="%.1f%%", min_value=0, max_value=max(float(tb["Taxa%"].max()) if not tb.empty else 1, 1))})

    if tem_vip_inf and not inf.empty:
        _sec("Detalhamento — Infância (VIP)")
        cols_det = [c for c in ["Número SA","FSLOI_GPONAccess","CODIGO_TECNICO_EXTRAIDO","NOME_TEC",
                                "Logradouro","Número","Bairro","Cidade",
                                "inf_dias_anterior","inf_tecnico_filho","inf_dat_fech_anterior"] if c in inf.columns]
        det = inf[cols_det].drop_duplicates(subset=["FSLOI_GPONAccess"])
        det.rename(columns={
            "FSLOI_GPONAccess":"GPON","Número SA":"SA Inst.",
            "CODIGO_TECNICO_EXTRAIDO":"TR (Instalador)",
            "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro","Cidade":"Cidade",
            "inf_dias_anterior":"Dias p/ reparo",
            "inf_tecnico_filho":"Tec. que Reparou"}, inplace=True)
        st.dataframe(det, use_container_width=True, hide_index=True)

        st.divider()

        _sec("Infância por Dia do Mês")
        if "DIA_FIM" in inf.columns:
            diario = inf.groupby("DIA_FIM").size().reset_index(name="Infancia")
            diario["Dia"] = pd.to_datetime(diario["DIA_FIM"]).dt.strftime("%d/%m")
            fig_dia = make_subplots(specs=[[{"secondary_y": False}]])
            fig_dia.add_bar(x=diario["Dia"], y=diario["Infancia"], name="Infância", marker_color=ROYAL)
            fig_dia.add_scatter(x=diario["Dia"], y=diario["Infancia"], mode="lines+markers",
                                line=dict(color=ROYAL_LIGHT, width=2), marker=dict(size=6, color=ROYAL_LIGHT),
                                name="Tendência")
            fig_dia.update_layout(**_lyt("", 300))
            fig_dia.update_yaxes(title_text="Quantidade")
            st.plotly_chart(fig_dia, use_container_width=True)
        else:
            st.info("Dados insuficientes para gráfico diário.")

        _sec("Pareto de Causas da Infância")
        sas = inf["SA_REPARO_INFANCIA"].dropna().unique() if "SA_REPARO_INFANCIA" in inf.columns else []
        rows = ds[ds["Número SA"].isin(sas)] if len(sas) else pd.DataFrame()
        if not rows.empty:
            mapa_causas = carregar_causas_macro()
            if mapa_causas:
                rows["CAUSA_MACRO"] = rows["Código de encerramento"].astype(str).str.strip().map(mapa_causas)
                rows["CAUSA_MACRO"] = rows["CAUSA_MACRO"].fillna("OUTROS")
            else:
                rows["CAUSA_MACRO"] = rows["Descrição"].str[:50]
            causas = rows.groupby("CAUSA_MACRO").size().reset_index(name="Qtd")
            causas = causas.sort_values("Qtd", ascending=False).head(15)
            causas["Perc"] = (causas["Qtd"] / causas["Qtd"].sum() * 100).round(1)
            causas["Acum"] = causas["Perc"].cumsum()
            fig_causas = make_subplots(specs=[[{"secondary_y": True}]])
            fig_causas.add_bar(x=causas["CAUSA_MACRO"], y=causas["Qtd"],
                               name="Ocorrências", marker_color=ROYAL)
            fig_causas.add_scatter(x=causas["CAUSA_MACRO"], y=causas["Acum"],
                                   name="% Acumulado", mode="lines+markers",
                                   line=dict(color=ROYAL_LIGHT, width=2), secondary_y=True)
            fig_causas.update_layout(**_lyt("Pareto de Causas", 350))
            fig_causas.update_yaxes(showgrid=False, secondary_y=True)
            st.plotly_chart(fig_causas, use_container_width=True)
        else:
            st.info("Sem dados de causa dos reparos de infância.")

        _sec("Pareto de Técnicos — Infância")
        pareto_tec = tb[["Nome","TR","Infancia"]].sort_values("Infancia", ascending=False)
        pareto_tec["Perc"] = (pareto_tec["Infancia"] / pareto_tec["Infancia"].sum() * 100).round(1)
        pareto_tec["Acum"] = pareto_tec["Perc"].cumsum()
        fig_tec = make_subplots(specs=[[{"secondary_y": True}]])
        fig_tec.add_bar(x=pareto_tec["Nome"], y=pareto_tec["Infancia"],
                        name="Infância", marker_color=ROYAL)
        fig_tec.add_scatter(x=pareto_tec["Nome"], y=pareto_tec["Acum"],
                            name="% Acumulado", mode="lines+markers",
                            line=dict(color=ROYAL_LIGHT, width=2), secondary_y=True)
        fig_tec.update_layout(**_lyt("Pareto de Técnicos", 400))
        fig_tec.update_yaxes(showgrid=False, secondary_y=True)
        st.plotly_chart(fig_tec, use_container_width=True)


def tela_calendario(df, ds, f):
    import calendar as _cal
    _header("📆", "Calendario Mensal", f)

=======
def tela_calendario(df, f):
    import calendar as _cal
    _header("📆", "Calendário Mensal", f)
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)
    mes_str = f["mes"]
    try:
        per = pd.Period(mes_str, freq="M")
        ano, mes = per.year, per.month
    except Exception:
<<<<<<< HEAD
        st.error("Mes invalido."); return

    hoje      = datetime.now()
    dias_mes  = _cal.monthrange(ano, mes)[1]
    dia_max   = hoje.day if (ano == hoje.year and mes == hoje.month) else dias_mes
    dias_range = list(range(1, dia_max + 1))
    meses_pt  = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
    lbl_mes   = f"{meses_pt[mes-1]}/{ano}"

    df_m = ds[(ds["FIM_DT"].dt.year==ano) & (ds["FIM_DT"].dt.month==mes)].copy()
    df_m["DIA"] = df_m["FIM_DT"].dt.day.astype(int)
    suc_m    = df_m[df_m["FLAG_CONCLUIDO_SUCESSO"]    == "SIM"]
    semsuc_m = df_m[df_m["FLAG_CONCLUIDO_SEM_SUCESSO"] == "SIM"]

    if suc_m.empty:
        st.warning("Sem dados de producao para o mes selecionado.")
        return

    tecs = suc_m["CODIGO_TECNICO_EXTRAIDO"].nunique()
    efic = round(len(suc_m)/(len(suc_m)+len(semsuc_m))*100,1) if (len(suc_m)+len(semsuc_m))>0 else 0
    cols = st.columns(5)
    for col,(lb,vl,sb,cl) in zip(cols,[
        ("Tecnicos Ativos", f"{tecs}",           lbl_mes,       "kpi-blue"),
        ("Concluidos",      f"{len(suc_m):,}",   "c/ sucesso",  "kpi-blue"),
        ("INST",            f"{(suc_m['Macro Atividade']=='INST-FTTH').sum():,}", "", "kpi-blue"),
        ("REP",             f"{(suc_m['Macro Atividade']=='REP-FTTH').sum():,}",  "", "kpi-purple"),
        ("Eficacia",        f"{efic}%",           "suc/total",   "kpi-green" if efic>=85 else "kpi-yellow"),
    ]):
        col.markdown(_kpi(lb,vl,sb,cl), unsafe_allow_html=True)
    st.write("")

    st.markdown("""
    <div style="display:flex;gap:10px;margin-bottom:10px;font-size:11px;align-items:center;">
        <span style="background:#7c3aed;color:white;padding:2px 8px;border-radius:4px;">≥8</span>
        <span style="background:#000000;color:white;padding:2px 8px;border-radius:4px;">6-7</span>
        <span style="background:#1e40af;color:white;padding:2px 8px;border-radius:4px;">5</span>
        <span style="background:#d4edda;color:#155724;padding:2px 8px;border-radius:4px;">4</span>
        <span style="background:#fff3cd;color:#856404;padding:2px 8px;border-radius:4px;">3</span>
        <span style="background:#ffcccc;color:#721c24;padding:2px 8px;border-radius:4px;">1-2</span>
        <span style="background:#f0f0f0;color:#6c757d;padding:2px 8px;border-radius:4px;">0</span>
        <span style="color:#64748b;margin-left:4px;">= atividades no dia</span>
    </div>
    """, unsafe_allow_html=True)

    _sec(f"Producao por Tecnico — {lbl_mes}")

    def _n(v): return str(v).split(" - ")[0].strip().title() if pd.notna(v) else ""

    suc_range = suc_m[suc_m["DIA"].isin(dias_range)]
    pivot = suc_range.groupby(["CODIGO_TECNICO_EXTRAIDO","DIA"]).size().unstack(fill_value=0)
    for d in dias_range:
        if d not in pivot.columns:
            pivot[d] = 0
    pivot = pivot[dias_range]

    nomes     = suc_m.groupby("CODIGO_TECNICO_EXTRAIDO")["Técnico Atribuído"].first().apply(_n)
    ss_tec    = semsuc_m.groupby("CODIGO_TECNICO_EXTRAIDO").size()
    dias_trab = suc_range.groupby("CODIGO_TECNICO_EXTRAIDO")["DIA"].nunique()

    pivot.insert(0, "Nome", nomes)
    pivot["Total"]     = pivot[dias_range].sum(axis=1)
    pivot["Sem Suc."]  = ss_tec.reindex(pivot.index).fillna(0).astype(int)
    pivot["Dias"]      = dias_trab.reindex(pivot.index).fillna(0).astype(int)
    pivot["Media"]     = (pivot["Total"] / pivot["Dias"].replace(0,1)).round(1)
    pivot["Eficacia%"] = (pivot["Total"]/(pivot["Total"]+pivot["Sem Suc."]).replace(0,1)*100).round(1)
    pivot = pivot.sort_values("Total", ascending=False).reset_index()
    pivot = pivot.rename(columns={"CODIGO_TECNICO_EXTRAIDO":"TR"})

    cols_dia  = [str(d) for d in dias_range]
    cols_order = ["Nome","TR"] + dias_range + ["Dias","Total","Sem Suc.","Media","Eficacia%"]
    pivot = pivot[cols_order]
    pivot.columns = ["Nome","TR"] + cols_dia + ["Dias","Total","Sem Suc.","Media","Eficacia%"]

    def _cor_cel(v):
        try:
            v = int(v)
        except: return ""
        if v >= 8: return "background-color:#7c3aed;color:white;font-weight:700;text-align:center"
        if v in (6,7): return "background-color:#000000;color:white;font-weight:700;text-align:center"
        if v == 5: return "background-color:#1e40af;color:white;font-weight:700;text-align:center"
        if v == 4: return "background-color:#d4edda;color:#155724;font-weight:600;text-align:center"
        if v == 3: return "background-color:#fff3cd;color:#856404;font-weight:600;text-align:center"
        if v in (1,2): return "background-color:#ffcccc;color:#721c24;text-align:center"
        return "background-color:#f0f0f0;color:#adb5bd;text-align:center"

    try:
        styled = pivot.style.map(_cor_cel, subset=cols_dia)
    except AttributeError:
        styled = pivot.style.applymap(_cor_cel, subset=cols_dia)

    max_p = int(pivot["Total"].max()) if not pivot.empty else 1
    altura_total = 38 + (len(pivot) * 35)

    st.dataframe(styled, use_container_width=True, hide_index=True,
        column_config={
            "Nome"     : st.column_config.TextColumn("Nome", width="medium"),
            "TR"       : st.column_config.TextColumn("TR", width="small"),
            "Total"    : st.column_config.ProgressColumn("Total", format="%d", min_value=0, max_value=max_p),
            "Eficacia%": st.column_config.ProgressColumn("Eficacia%", format="%.1f%%", min_value=0, max_value=100),
            "Media"    : st.column_config.NumberColumn("Media", format="%.1f"),
        }, height=altura_total)

    ef_list = []
    for d in dias_range:
        rows = df_m[df_m["DIA"] == d]
        s = (rows["FLAG_CONCLUIDO_SUCESSO"]=="SIM").sum()
        ss = (rows["FLAG_CONCLUIDO_SEM_SUCESSO"]=="SIM").sum()
        ef_list.append({"DIA": d, "Concluidos": s, "Eficacia%": round(s / max(s+ss, 1) * 100, 1)})
    ef_dia = pd.DataFrame(ef_list)

    fig = go.Figure()
    fig.add_bar(x=ef_dia["DIA"].astype(str), y=ef_dia["Concluidos"], name="Concluidos", marker_color=ROYAL, yaxis="y")
    fig.add_scatter(x=ef_dia["DIA"].astype(str), y=ef_dia["Eficacia%"], name="Eficacia%", mode="lines+markers",
                    line=dict(color=ROYAL_LIGHT,width=2), marker_size=6, yaxis="y2")
    fig.add_hline(y=85, line_dash="dash", line_color=C["yellow"], annotation_text="Meta 85%", yref="y2")
    fig.update_layout(
        **_lyt(f"Producao e Eficacia — {lbl_mes}", 320),
        yaxis2=dict(overlaying="y", side="right", showgrid=False, tickfont_color=C["green"], range=[0,110]),
        showlegend=True)
    st.plotly_chart(fig, use_container_width=True)


def tela_diario(df, ds, f):
    _header("📅", "Controle do Dia", f)

    datas_disp = sorted(ds["FIM_DT"].dropna().dt.date.unique(), reverse=True)
    if not datas_disp:
        st.warning("Nenhuma data disponivel.")
        return

    hoje = datetime.today().date()
    if hoje in datas_disp:
        valor_padrao = hoje
    else:
        valor_padrao = datas_disp[0]

    c_pick, c_info = st.columns([2, 5])
    with c_pick:
        dia_sel = st.date_input("Data de referencia",
            value=valor_padrao,
            min_value=datas_disp[-1],
            max_value=datas_disp[0],
            key="dia_ref")

    dm  = ds[ds["FIM_DT"].dt.date == dia_sel].copy()
    dm_ab = ds[ds["AB_DT"].dt.date == dia_sel].copy()

    suc     = dm[dm["FLAG_CONCLUIDO_SUCESSO"]    == "SIM"]
    sem_suc = dm[dm["FLAG_CONCLUIDO_SEM_SUCESSO"] == "SIM"]
    inst_d  = suc[suc["Macro Atividade"] == "INST-FTTH"]
    rep_d   = suc[suc["Macro Atividade"] == "REP-FTTH"]
    tecs_atv = suc["CODIGO_TECNICO_EXTRAIDO"].nunique()
    efic    = round(len(suc)/(len(suc)+len(sem_suc))*100,1) if (len(suc)+len(sem_suc))>0 else 0

    with c_info:
        st.markdown(
            f"<div style='margin-top:28px;font-size:12px;color:#64748b;'>"
            f"<b>{dia_sel.strftime('%d/%m/%Y')}</b> — "
            f"{tecs_atv} tecnicos ativos | {len(suc)+len(sem_suc)} atividades concluidas</div>",
            unsafe_allow_html=True)

    rep_dia_ab = dm_ab[dm_ab["FLAG_REPETIDO_30D"]  == "SIM"]
    rep_ab_tot = ds[ds["FLAG_REPETIDO_ABERTO"]      == "SIM"]
    inf_dia    = suc[suc["FLAG_INFANCIA_30D"]        == "SIM"]
    p0_10 = ds[(ds["FIM_DT"].dt.date == dia_sel) & (ds["FLAG_P0_10_DIA"] == "SIM")]
    p0_15 = ds[(ds["FIM_DT"].dt.date == dia_sel) & (ds["FLAG_P0_15_DIA"] == "SIM")]

    cols = st.columns(7)
    for col, (lb,vl,sb,cl) in zip(cols, [
        ("Concluidos",    f"{len(suc):,}",  f"INST:{len(inst_d)} REP:{len(rep_d)}", "kpi-blue"),
        ("Eficacia",      f"{efic}%",        "suc/total",   "kpi-green" if efic>=85 else "kpi-yellow" if efic>=70 else "kpi-red"),
        ("Sem Sucesso",   f"{len(sem_suc):,}","pendencias",  "kpi-red" if len(sem_suc)>0 else "kpi-green"),
        ("Rep. Dia",      f"{len(rep_dia_ab):,}","abertos hoje","kpi-red" if len(rep_dia_ab)>0 else "kpi-green"),
        ("Rep. Abertos",  f"{len(rep_ab_tot):,}","em garantia","kpi-yellow" if len(rep_ab_tot)>0 else "kpi-green"),
        ("P0 10h",        f"{p0_10['CODIGO_TECNICO_EXTRAIDO'].nunique()}","tecnicos","kpi-red" if not p0_10.empty else "kpi-green"),
        ("P0 15h",        f"{p0_15['CODIGO_TECNICO_EXTRAIDO'].nunique()}","tecnicos","kpi-red" if not p0_15.empty else "kpi-green"),
    ]):
        col.markdown(_kpi(lb,vl,sb,cl), unsafe_allow_html=True)
    st.write("")

    def _extrair_tr(nome):
        if pd.isna(nome) or not nome: return ""
        m = re.search(r"(TR|TT|TC)\d+", str(nome), re.I)
        return m.group(0).upper() if m else ""

    _sec("Produtividade por Tecnico")
    def _n(v): return str(v).split(" - ")[0].strip().title() if pd.notna(v) else ""
    if suc.empty:
        st.info("Nenhuma atividade concluida neste dia.")
    else:
        pt = suc.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
            Nome  =("Técnico Atribuído", lambda x: _n(x.iloc[0])),
            Total =("Número SA","count"),
            INST  =("Macro Atividade", lambda x: (x=="INST-FTTH").sum()),
            REP   =("Macro Atividade", lambda x: (x=="REP-FTTH").sum()),
        ).reset_index()
        ss_tec = sem_suc.groupby("CODIGO_TECNICO_EXTRAIDO").size().reset_index(name="SemSuc")
        pt = pt.merge(ss_tec, on="CODIGO_TECNICO_EXTRAIDO", how="left")
        pt["SemSuc"] = pt["SemSuc"].fillna(0).astype(int)
        pt["Efic%"]  = (pt["Total"]/(pt["Total"]+pt["SemSuc"]).replace(0,1)*100).round(1)
        pt = pt.sort_values("Total", ascending=False).reset_index(drop=True)
        pt.columns = ["TR","Nome","Total","INST","REP","Sem Suc.","Eficacia%"]
        st.dataframe(pt, use_container_width=True, hide_index=True,
            column_config={
                "Total":    st.column_config.ProgressColumn("Total", format="%d", min_value=0,
                            max_value=int(pt["Total"].max()) if not pt.empty else 1),
                "Eficacia%":st.column_config.ProgressColumn("Eficacia%", format="%.1f%%", min_value=0, max_value=100),
            })

    _sec("Sem Sucesso — Pendencias do Dia")
    if sem_suc.empty:
        st.success("Nenhuma pendencia no dia.")
    else:
        ok = [c for c in ["Número SA","CODIGO_TECNICO_EXTRAIDO","NOME_TEC",
                           "Macro Atividade","Descrição","Observação",
                           "Logradouro","Número","Bairro","Cidade",
                           "Código de encerramento"] if c in sem_suc.columns]
        st.dataframe(sem_suc[ok].rename(columns={
            "Número SA":"SA","CODIGO_TECNICO_EXTRAIDO":"TR","NOME_TEC":"Tecnico",
            "Macro Atividade":"Tipo","Código de encerramento":"Cod. Enc.",
            "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro","Cidade":"Cidade"}),
            use_container_width=True, hide_index=True)

    c1, c2 = st.columns(2)
    with c1:
        _sec("Repetidos Abertos no Dia")
        if rep_dia_ab.empty:
            st.success("Nenhum repetido aberto hoje.")
        else:
            cols_rep = ["Número SA","CODIGO_TECNICO_EXTRAIDO","NOME_TEC","FSLOI_GPONAccess","Logradouro","Número","Bairro"]
            if "rep_tecnico_pai" in rep_dia_ab.columns:
                rep_dia_ab["TR_PAI"] = rep_dia_ab["rep_tecnico_pai"].apply(_extrair_tr)
                cols_rep.extend(["rep_tecnico_pai","TR_PAI"])
            if "ALARMADO" in rep_dia_ab.columns:
                cols_rep.append("ALARMADO")
            df_show = rep_dia_ab[[c for c in cols_rep if c in rep_dia_ab.columns]].copy()
            rename_map = {
                "Número SA":"SA","CODIGO_TECNICO_EXTRAIDO":"TR",
                "NOME_TEC":"Tecnico","FSLOI_GPONAccess":"GPON",
                "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro",
                "rep_tecnico_pai":"Tecnico (PAI)","TR_PAI":"TR (PAI)"
            }
            st.dataframe(df_show.rename(columns=rename_map), use_container_width=True, hide_index=True)

    with c2:
        _sec("Repetidos em Garantia (Abertos)")
        if rep_ab_tot.empty:
            st.success("Nenhum reparo em garantia.")
        else:
            cols_rep = ["Número SA","CODIGO_TECNICO_EXTRAIDO","NOME_TEC","FSLOI_GPONAccess","Logradouro","Número","Bairro","DIA_AB"]
            if "rep_tecnico_pai" in rep_ab_tot.columns:
                rep_ab_tot["TR_PAI"] = rep_ab_tot["rep_tecnico_pai"].apply(_extrair_tr)
                cols_rep.extend(["rep_tecnico_pai","TR_PAI"])
            if "ALARMADO" in rep_ab_tot.columns:
                cols_rep.append("ALARMADO")
            df_show = rep_ab_tot[[c for c in cols_rep if c in rep_ab_tot.columns]].copy()
            rename_map = {
                "Número SA":"SA","CODIGO_TECNICO_EXTRAIDO":"TR",
                "NOME_TEC":"Tecnico","FSLOI_GPONAccess":"GPON","DIA_AB":"Abertura",
                "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro",
                "rep_tecnico_pai":"Tecnico (PAI)","TR_PAI":"TR (PAI)"
            }
            st.dataframe(df_show.rename(columns=rename_map), use_container_width=True, hide_index=True)

    c3, c4 = st.columns(2)
    with c3:
        _sec("Infancia — Instalacoes do Dia")
        if inf_dia.empty:
            st.success("Nenhuma infancia hoje.")
        else:
            cols_inf = ["Número SA","CODIGO_TECNICO_EXTRAIDO","NOME_TEC","FSLOI_GPONAccess","Logradouro","Número","Bairro"]
            if "SA_REPARO_INFANCIA" in inf_dia.columns:
                cols_inf.append("SA_REPARO_INFANCIA")
            if "inf_tecnico_pai" in inf_dia.columns:
                inf_dia["TR_PAI"] = inf_dia["inf_tecnico_pai"].apply(_extrair_tr)
                cols_inf.extend(["inf_tecnico_pai","TR_PAI"])
            df_show = inf_dia[[c for c in cols_inf if c in inf_dia.columns]].copy()
            rename_map = {
                "Número SA":"SA Inst.","CODIGO_TECNICO_EXTRAIDO":"TR",
                "NOME_TEC":"Tecnico","FSLOI_GPONAccess":"GPON",
                "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro",
                "SA_REPARO_INFANCIA":"SA Reparo",
                "inf_tecnico_pai":"Tecnico (PAI)","TR_PAI":"TR (PAI)"
            }
            st.dataframe(df_show.rename(columns=rename_map), use_container_width=True, hide_index=True)

    with c4:
        _sec("Infancia Aberta (Reparo em Andamento)")
        estados_ab = ["ATRIBUÍDO","NÃO ATRIBUÍDO","RECEBIDO","EM EXECUÇÃO","EM DESLOCAMENTO"]
        gpons_suc = set(suc["FSLOI_GPONAccess"].dropna().str.upper())
        inf_ab_dia = ds[
            (ds["Macro Atividade"] == "REP-FTTH") &
            (ds["Estado"].isin(estados_ab)) &
            (ds["FLAG_INFANCIA_30D"] == "SIM") &
            (ds["FSLOI_GPONAccess"].str.upper().isin(gpons_suc))
        ].copy()
        if inf_ab_dia.empty:
            st.success("Nenhuma infancia aberta.")
        else:
            cols_ab = ["Número SA","CODIGO_TECNICO_EXTRAIDO","NOME_TEC","FSLOI_GPONAccess","Estado","Logradouro","Número","Bairro"]
            if "inf_tecnico_pai" in inf_ab_dia.columns:
                inf_ab_dia["TR_PAI"] = inf_ab_dia["inf_tecnico_pai"].apply(_extrair_tr)
                cols_ab.extend(["inf_tecnico_pai","TR_PAI"])
            df_show = inf_ab_dia[[c for c in cols_ab if c in inf_ab_dia.columns]].copy()
            rename_map = {
                "Número SA":"SA","CODIGO_TECNICO_EXTRAIDO":"TR",
                "NOME_TEC":"Tecnico","FSLOI_GPONAccess":"GPON",
                "Logradouro":"Endereço","Número":"Nº","Bairro":"Bairro",
                "inf_tecnico_pai":"Tecnico (PAI)","TR_PAI":"TR (PAI)"
            }
            st.dataframe(df_show.rename(columns=rename_map), use_container_width=True, hide_index=True)

    _sec("P0 — Controle de Encerramento")
    cp1, cp2 = st.columns(2)
    with cp1:
        st.markdown("**P0 10h — Nao encerraram ate as 10h**")
        if p0_10.empty:
            st.success("Todos encerraram ate 10h.")
        else:
            t10 = p0_10.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
                Nome=("Técnico Atribuído", lambda x: _n(x.iloc[0])),
                Qtd =("Número SA","count")).reset_index()
            t10.columns = ["TR","Nome","Qtd"]
            st.dataframe(t10, use_container_width=True, hide_index=True)
    with cp2:
        st.markdown("**P0 15h — Nao encerraram ate as 15h**")
        if p0_15.empty:
            st.success("Todos encerraram ate 15h.")
        else:
            t15 = p0_15.groupby("CODIGO_TECNICO_EXTRAIDO").agg(
                Nome=("Técnico Atribuído", lambda x: _n(x.iloc[0])),
                Qtd =("Número SA","count")).reset_index()
            t15.columns = ["TR","Nome","Qtd"]
            st.dataframe(t15, use_container_width=True, hide_index=True)


# =============================================================================
# 8. TELA QUALIDADE
# =============================================================================

def _cls_prod(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ("S/D", "⬜", "#94a3b8", 0)
    v = float(v)
    if v >= 6:   return ("EXCELENTE",        "🏆", "#1e3a5f", 4)
    if v >= 5:   return ("PARABENS",         "🟢", "#16a34a", 3)
    if v >= 4:   return ("PARABENS",         "🟢", "#16a34a", 3)
    if v >= 3:   return ("ATENCAO",          "🟡", "#d97706", 2)
    if v >= 1:   return ("PRECISA MELHORAR", "🔴", "#dc2626", 1)
    return            ("PRECISA MELHORAR",   "🔴", "#dc2626", 1)

def _cls_efic(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ("S/D", "⬜", "#94a3b8", 0)
    v = float(v)
    if v >= 85:  return ("EXCELENTE",        "🏆", "#1e3a5f", 4)
    if v >= 82:  return ("PARABENS",         "🟢", "#16a34a", 3)
    if v >= 75:  return ("ATENCAO",          "🟡", "#d97706", 2)
    return            ("PRECISA MELHORAR",   "🔴", "#dc2626", 1)

def _cls_rep(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ("S/D", "⬜", "#94a3b8", 0)
    v = float(v)
    if v < 2:    return ("EXCELENTE",        "🏆", "#1e3a5f", 4)
    if v <= 5:   return ("OTIMO",            "🟢", "#16a34a", 3)
    if v <= 9:   return ("PARABENS",         "🟢", "#16a34a", 2)
    return            ("PRECISA MELHORAR",   "🔴", "#dc2626", 0)

def _cls_inf(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ("S/D", "⬜", "#94a3b8", 0)
    v = float(v)
    if v < 3:    return ("OTIMO",            "🟢", "#16a34a", 3)
    return            ("PRECISA MELHORAR",   "🔴", "#dc2626", 0)

def _nota_total(pc, ec, rc, ic):
    return pc[3] + ec[3] + rc[3] + ic[3]

def _cor_nota(n):
    if n >= 13: return "#1e3a5f"
    if n >= 9:  return "#16a34a"
    if n >= 5:  return "#d97706"
    return "#dc2626"

def _badge(label, cor):
    return (f'<span style="background:{cor};color:white;padding:2px 10px;'
            f'border-radius:12px;font-size:11px;font-weight:700">{label}</span>')

def _html_ata_qualidade(nome, codigo, supervisor, mes_ref,
                        prod_val, efic_val, rep_val, inf_val,
                        prod_cls, efic_cls, rep_cls, inf_cls, nota):
    nota_max = 16
    titulo = "ATA DE DESEMPENHO OPERACIONAL"
    cor_h = "135deg,#1e3a5f 0%,#2563eb 100%"
    
    if nota >= 13:
        msg = (f"Prezado(a) <strong>{nome}</strong>, seu desempenho no periodo foi "
               f"excepcional. Todos os indicadores estao dentro ou acima das metas. Continue assim!")
    elif nota >= 9:
        msg = (f"Prezado(a) <strong>{nome}</strong>, seu desempenho esta dentro das metas "
               f"esperadas. Identifique os pontos de atencao para atingir a excelencia.")
    elif nota >= 5:
        msg = (f"Prezado(a) <strong>{nome}</strong>, ha indicadores que precisam de atencao. "
               f"Vamos alinhar um plano de acao para a melhoria do desempenho.")
    else:
        msg = (f"Prezado(a) <strong>{nome}</strong>, seus indicadores estao abaixo das metas "
               f"estabelecidas. E necessario um plano de acao imediato.")

    def _ind(lbl, val, cls):
        return (f'<div style="flex:1;min-width:140px;border:1px solid #e2e8f0;border-radius:8px;'
                f'padding:14px;text-align:center">'
                f'<div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">{lbl}</div>'
                f'<div style="font-size:20px;font-weight:700;color:#1a3a8f;margin-bottom:8px">{val}</div>'
                f'{_badge(cls[0], cls[2])}'
                f'</div>')

    inds_html = (
        _ind("Produtividade", prod_val, prod_cls) +
        _ind("Eficacia",      efic_val, efic_cls) +
        _ind("Repetida",      rep_val,  rep_cls)  +
        _ind("Infancia",      inf_val,  inf_cls)
    )

    cod_safe = codigo.replace("'","\\'")
    nome_safe = nome.replace("'","\\'")
    sup_safe  = supervisor.replace("'","\\'")
    ts_now    = datetime.now().strftime("%Y%m%d_%H%M%S")
    dt_now    = datetime.now().strftime("%d/%m/%Y %H:%M")

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Ata — {nome}</title>
<script src="https://cdn.jsdelivr.net/npm/signature_pad@4.0.0/dist/signature_pad.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jspdf/2.5.1/jspdf.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box;font-family:'Segoe UI',sans-serif}}
body{{background:#f5f7fa;padding:20px;color:#333;line-height:1.6}}
.box{{max-width:860px;margin:0 auto;background:white;border-radius:12px;
      box-shadow:0 5px 25px rgba(0,0,0,.08);overflow:hidden}}
.hdr{{background:linear-gradient({cor_h});color:white;padding:24px 30px;text-align:center}}
.hdr h1{{font-size:20px;margin-bottom:4px;font-weight:700}}
.hdr p{{font-size:12px;opacity:.9}}
.body{{padding:26px}}
.info{{background:#f0f4ff;border-radius:8px;padding:14px;margin-bottom:18px;
       border-left:4px solid #1a3a8f;font-size:13px}}
.info p{{margin:2px 0}}
.msg{{background:#f8fafc;padding:13px;border-radius:8px;font-style:italic;
      margin:16px 0;border-left:4px solid #1a3a8f;font-size:13px}}
.inds{{display:flex;gap:12px;flex-wrap:wrap;margin:16px 0}}
.nota{{text-align:center;margin:16px 0}}
.nota span{{font-size:28px;font-weight:800;color:#1a3a8f}}
.sig-box{{background:#f9fafc;border-radius:8px;padding:14px;
          border:1px solid #eaeaea;margin-top:14px}}
.sig-title{{font-size:13px;color:#1a3a8f;margin-bottom:8px;font-weight:600}}
.sig-area{{position:relative;width:100%;height:120px;border:2px dashed #bdc3c7;
           border-radius:8px;background:white;margin-bottom:10px}}
.sig-canvas{{width:100%;height:100%;display:block;cursor:crosshair}}
.sig-ph{{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);
         color:#95a5a6;font-size:13px;text-align:center;pointer-events:none}}
.btns{{display:flex;gap:8px;margin-bottom:8px}}
.btn{{padding:8px 14px;border:none;border-radius:6px;font-weight:600;
      cursor:pointer;font-size:13px;flex:1}}
.bp{{background:#1a3a8f;color:white}}
.bs{{background:#ecf0f1;color:#333}}
.sig-info{{background:#e8f4fc;border-radius:8px;padding:10px;margin-top:6px;
           border-left:4px solid #1a3a8f;display:none;font-size:12px}}
.sig-info.on{{display:block}}
.btn-pdf{{background:#16a34a;color:white;border:none;border-radius:6px;
          padding:12px;font-size:15px;font-weight:700;cursor:pointer;
          width:100%;margin-top:18px}}
.rodape{{margin-top:16px;padding-top:12px;border-top:1px solid #eaeaea;
         font-size:11px;color:#94a3b8;text-align:center}}
@media(max-width:560px){{.inds{{flex-direction:column}}}}
</style>
</head>
<body>
<div class="box">
  <div class="hdr">
    <h1>{titulo}</h1>
    <p>Periodo: {mes_ref} &nbsp;·&nbsp; Gerado em: {dt_now}</p>
  </div>
  <div class="body">
    <div class="info">
      <p><strong>Nome:</strong> {nome}</p>
      <p><strong>Codigo:</strong> {codigo}</p>
      <p><strong>Supervisor:</strong> {supervisor}</p>
    </div>
    <div class="msg">{msg}</div>

    <h3 style="color:#1a3a8f;font-size:14px;margin-bottom:10px">📊 Indicadores do Periodo</h3>
    <div class="inds">{inds_html}</div>
    <div class="nota">Pontuacao: <span>{nota}/{nota_max}</span></div>

    <!-- Assinatura Tecnico -->
    <div class="sig-box">
      <div class="sig-title">✍️ Assinatura do Tecnico <span style="color:#ef4444">*</span></div>
      <div class="sig-area">
        <canvas id="cvT" class="sig-canvas"></canvas>
        <div id="phT" class="sig-ph">Assine aqui (obrigatorio)</div>
      </div>
      <div class="btns">
        <button class="btn bs" id="clrT">🧹 Limpar</button>
        <button class="btn bp" id="valT">✔ Validar</button>
      </div>
      <div id="infoT" class="sig-info">
        <strong>Tecnico:</strong> {nome} &nbsp;|&nbsp;
        <strong>Data/Hora:</strong> <span id="dtT">--</span>
      </div>
    </div>

    <!-- Assinatura Supervisor -->
    <div class="sig-box" style="margin-top:14px">
      <div class="sig-title">✍️ Assinatura do Supervisor <span style="color:#94a3b8">(opcional)</span></div>
      <div class="sig-area">
        <canvas id="cvS" class="sig-canvas"></canvas>
        <div id="phS" class="sig-ph">Assine aqui (opcional)</div>
      </div>
      <div class="btns">
        <button class="btn bs" id="clrS">🧹 Limpar</button>
        <button class="btn bp" id="valS">✔ Validar</button>
      </div>
      <div id="infoS" class="sig-info">
        <strong>Supervisor:</strong> {supervisor} &nbsp;|&nbsp;
        <strong>Data/Hora:</strong> <span id="dtS">--</span>
      </div>
    </div>

    <button class="btn-pdf" id="btnPDF" disabled>📄 Salvar Ata e Gerar PDF</button>
    <div class="rodape">BERTA — Painel Operacional FTTH/GPON — Santa Catarina</div>
  </div>
</div>
<script>
let padT, padS, tecOk=false;
function resize(cv){{const r=cv.getBoundingClientRect();cv.width=r.width;cv.height=r.height;}}
window.onload=()=>{{
  const cvT=document.getElementById('cvT'), cvS=document.getElementById('cvS');
  resize(cvT); resize(cvS);
  padT=new SignaturePad(cvT); padS=new SignaturePad(cvS);
  padT.addEventListener('beginStroke',()=>document.getElementById('phT').style.display='none');
  padS.addEventListener('beginStroke',()=>document.getElementById('phS').style.display='none');
}};
document.getElementById('clrT').onclick=()=>{{padT.clear();document.getElementById('phT').style.display='';tecOk=false;document.getElementById('btnPDF').disabled=true;document.getElementById('infoT').classList.remove('on');}};
document.getElementById('clrS').onclick=()=>{{padS.clear();document.getElementById('phS').style.display='';document.getElementById('infoS').classList.remove('on');}};
document.getElementById('valT').onclick=()=>{{
  if(padT.isEmpty()){{alert('Assine antes de validar.');return;}}
  document.getElementById('dtT').textContent=new Date().toLocaleString('pt-BR');
  document.getElementById('infoT').classList.add('on');
  tecOk=true; document.getElementById('btnPDF').disabled=false;
}};
document.getElementById('valS').onclick=()=>{{
  if(padS.isEmpty()){{alert('Assine antes de validar.');return;}}
  document.getElementById('dtS').textContent=new Date().toLocaleString('pt-BR');
  document.getElementById('infoS').classList.add('on');
}};
document.getElementById('btnPDF').onclick=async()=>{{
  if(!tecOk){{alert('Assinatura do tecnico e obrigatoria.');return;}}
  const {{jsPDF}}=window.jspdf;
  const doc=new jsPDF({{unit:'mm',format:'a4'}});
  const W=doc.internal.pageSize.getWidth();
  doc.setFillColor(26,58,143); doc.rect(0,0,W,38,'F');
  doc.setTextColor(255,255,255); doc.setFontSize(13); doc.setFont(undefined,'bold');
  doc.text('{titulo}',W/2,16,{{align:'center'}});
  doc.setFontSize(9); doc.setFont(undefined,'normal');
  doc.text('Periodo: {mes_ref}  |  Gerado: {dt_now}',W/2,26,{{align:'center'}});
  let y=46; doc.setTextColor(51,51,51); doc.setFontSize(11);
  doc.setFont(undefined,'bold'); doc.text('DADOS DO TECNICO',14,y); y+=7;
  doc.setFont(undefined,'normal');
  ['Nome: {nome_safe}','Codigo: {cod_safe}','Supervisor: {sup_safe}'].forEach(t=>{{doc.text(t,14,y);y+=6;}});
  y+=4; doc.setFont(undefined,'bold'); doc.text('INDICADORES',14,y); y+=7;
  doc.setFont(undefined,'normal');
  ['Produtividade: {prod_val} — {prod_cls[0]}','Eficacia: {efic_val} — {efic_cls[0]}',
   'Repetida: {rep_val} — {rep_cls[0]}','Infancia: {inf_val} — {inf_cls[0]}',
   'Pontuacao Geral: {nota}/{nota_max}'].forEach(t=>{{doc.text(t,14,y);y+=6;}});
  y+=6;
  if(!padT.isEmpty()){{
    doc.setFont(undefined,'bold'); doc.text('ASSINATURA DO TECNICO',14,y); y+=6;
    doc.addImage(padT.toDataURL('image/png'),'PNG',14,y,80,28); y+=32;
    doc.setFont(undefined,'normal'); doc.setFontSize(9);
    doc.text('Validada em: '+new Date().toLocaleString('pt-BR'),14,y); y+=8;
  }}
  if(!padS.isEmpty()){{
    doc.setFontSize(11); doc.setFont(undefined,'bold'); doc.text('ASSINATURA DO SUPERVISOR',14,y); y+=6;
    doc.addImage(padS.toDataURL('image/png'),'PNG',14,y,80,28); y+=32;
  }}
  const pdfBase64 = doc.output('datauristring').split(',')[1];
  const urlAtas = '{URL_ATAS_DRIVE}';
  const payload = {{
    supervisor: '{sup_safe}',
    tecnico: '{cod_safe}',
    pdf_base64: pdfBase64
  }};
  fetch(urlAtas, {{ method: 'POST', mode: 'no-cors', headers: {{ 'Content-Type': 'application/json' }}, body: JSON.stringify(payload) }})
    .then(() => alert('ATA salva no Google Drive com sucesso!'))
    .catch(err => alert('Erro ao salvar no Drive: ' + err));
  doc.save('ata_qualidade_{cod_safe}_{ts_now}.pdf');
}};
window.addEventListener('resize',()=>{{
  setTimeout(()=>{{
    [padT,padS].forEach(p=>{{if(p&&p.canvas){{const r=p.canvas.getBoundingClientRect();const d=p.toData();p.canvas.width=r.width;p.canvas.height=r.height;if(d&&d.length)p.fromData(d);}}}});
  }},200);
}});
</script>
</body>
</html>"""


def tela_qualidade(dm, ds, f):
    _header("🏆", "Qualidade", f)

    def _n(v):
        return str(v).split(" - ")[0].strip().title() if pd.notna(v) else ""

    mes_ref = f.get("mes", "")

    tecs = sorted(dm["CODIGO_TECNICO_EXTRAIDO"].dropna().unique())
    if not tecs:
        st.warning("Nenhum tecnico encontrado para os filtros selecionados.")
        return

    rows = []
    for cod in tecs:
        df_t = dm[dm["CODIGO_TECNICO_EXTRAIDO"] == cod].copy()
        if df_t.empty:
            continue

        nome = _n(df_t["Técnico Atribuído"].dropna().iloc[0]) if not df_t["Técnico Atribuído"].dropna().empty else cod

        suc = df_t[df_t["FLAG_CONCLUIDO_SUCESSO"] == "SIM"]
        dias_unicos = suc["DIA_FIM"].dropna().nunique()
        prod_media  = round(len(suc) / dias_unicos, 1) if dias_unicos > 0 else None

        n_suc  = (df_t["FLAG_CONCLUIDO_SUCESSO"]     == "SIM").sum()
        n_ss   = (df_t["FLAG_CONCLUIDO_SEM_SUCESSO"] == "SIM").sum()
        efic   = round(n_suc / (n_suc + n_ss) * 100, 1) if (n_suc + n_ss) > 0 else None

        rep_den = (df_t["Macro Atividade"] == "REP-FTTH").sum()
        rep_num = ((df_t["Macro Atividade"] == "REP-FTTH") &
                   (df_t["FLAG_REPETIDO"] == "SIM")).sum()
        rep_pct = round(rep_num / rep_den * 100, 1) if rep_den > 0 else None

        inst_den = (df_t["FLAG_INSTALACAO_VALIDA"] == "SIM").sum()
        inst_num = ((df_t["FLAG_INSTALACAO_VALIDA"] == "SIM") &
                    (df_t["FLAG_INFANCIA"] == "SIM")).sum()
        inf_pct  = round(inst_num / inst_den * 100, 1) if inst_den > 0 else None

        pc = _cls_prod(prod_media)
        ec = _cls_efic(efic)
        rc = _cls_rep(rep_pct)
        ic = _cls_inf(inf_pct)
        nota = _nota_total(pc, ec, rc, ic)

        rows.append({
            "cod":        cod,
            "nome":       nome,
            "prod_media": prod_media,
            "efic_pct":   efic,
            "rep_pct":    rep_pct,
            "inf_pct":    inf_pct,
            "prod_cls":   pc,
            "efic_cls":   ec,
            "rep_cls":    rc,
            "inf_cls":    ic,
            "nota":       nota,
            "n_suc":      n_suc,
            "dias":       dias_unicos,
        })

    if not rows:
        st.warning("Sem dados suficientes para calcular indicadores.")
        return

    df_q = pd.DataFrame(rows).sort_values("nota", ascending=False).reset_index(drop=True)

    n_exc = (df_q["nota"] >= 13).sum()
    n_bom = ((df_q["nota"] >= 9) & (df_q["nota"] < 13)).sum()
    n_atc = ((df_q["nota"] >= 5) & (df_q["nota"] < 9)).sum()
    n_mel = (df_q["nota"] < 5).sum()

    cols_kpi = st.columns(5)
    for col, lbl, val, cls in zip(
        cols_kpi,
        ["Tecnicos",      "🏆 Excelente",  "✅ Bom",      "⚠️ Atencao",  "🔴 Melhoria"],
        [len(df_q),       n_exc,            n_bom,          n_atc,          n_mel],
        ["kpi-blue",      "kpi-blue",       "kpi-green",    "kpi-yellow",   "kpi-red"],
    ):
        col.markdown(_kpi(lbl, val, "", cls), unsafe_allow_html=True)

    st.write("")
    st.write("")
    with st.expander("📋 Regras de classificacao", expanded=False):
        st.markdown("""| Indicador | 🏆 Excelente | 🟢 Parabens / Otimo | 🟡 Atencao | 🔴 Precisa Melhorar |
|---|---|---|---|---|
| **Produtividade** (ativ/dia) | ≥ 6 | 4 ou 5 = Parabens | 3 | ≤ 2 |
| **Eficacia** (%) | ≥ 85% | 82–85% | 75–82% | < 75% |
| **Repetida** (%) | < 2% | 2–5% = Otimo · 5–9% = Parabens | — | > 9% |
| **Infancia** (%) | — | < 3% = Otimo | — | ≥ 3% |

**Pontuacao:** cada indicador vale 0–4 pontos. Total 0–16.
≥13 Excelente | 9–12 Bom | 5–8 Atencao | <5 Precisa Melhorar""")

    _sec("Ranking de Qualidade por Tecnico")

    disp = []
    for _, r in df_q.iterrows():
        pv = f"{r['prod_media']:.1f}" if r['prod_media'] is not None else "S/D"
        ev = f"{r['efic_pct']:.1f}%"  if r['efic_pct']   is not None else "S/D"
        rv = f"{r['rep_pct']:.1f}%"   if r['rep_pct']    is not None else "S/D"
        iv = f"{r['inf_pct']:.1f}%"   if r['inf_pct']    is not None else "S/D"
        disp.append({
            "Nome":       r["nome"],
            "TR":         r["cod"],
            "Prod.":      pv,
            "Prod_S":     r["prod_cls"][0],
            "Efic.":      ev,
            "Efic_S":     r["efic_cls"][0],
            "Repet.":     rv,
            "Rep_S":      r["rep_cls"][0],
            "Infan.":     iv,
            "Inf_S":      r["inf_cls"][0],
            "Nota":       f"{r['nota']}/16",
        })

    df_disp = pd.DataFrame(disp)

    _cor_map = {
        "EXCELENTE":        "background-color:#1e3a5f;color:white;font-weight:700",
        "PARABENS":         "background-color:#16a34a;color:white;font-weight:700",
        "OTIMO":            "background-color:#15803d;color:white;font-weight:700",
        "ATENCAO":          "background-color:#d97706;color:white;font-weight:700",
        "PRECISA MELHORAR": "background-color:#dc2626;color:white;font-weight:700",
        "S/D":              "background-color:#e2e8f0;color:#64748b",
    }
    def _cor(v):
        return _cor_map.get(v, "")

    _scols = ["Prod_S", "Efic_S", "Rep_S", "Inf_S"]
    try:
        styled = df_disp.style.applymap(_cor, subset=_scols)
    except AttributeError:
        styled = df_disp.style.map(_cor, subset=_scols)

    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        height=min(60 + len(df_disp) * 36, 700),
        column_config={
            "Nome":   st.column_config.TextColumn("Nome",   width="medium"),
            "TR":     st.column_config.TextColumn("TR",     width="small"),
            "Nota":   st.column_config.TextColumn("Nota",   width="small"),
            "Prod.":  st.column_config.TextColumn("Prod.ativ/dia", width="small"),
            "Efic.":  st.column_config.TextColumn("Eficacia",      width="small"),
            "Repet.": st.column_config.TextColumn("Repetida",      width="small"),
            "Infan.": st.column_config.TextColumn("Infancia",      width="small"),
            "Prod_S": st.column_config.TextColumn("Status Prod.",  width="medium"),
            "Efic_S": st.column_config.TextColumn("Status Efic.",  width="medium"),
            "Rep_S":  st.column_config.TextColumn("Status Rep.",   width="medium"),
            "Inf_S":  st.column_config.TextColumn("Status Inf.",   width="medium"),
        }
    )

    st.write("")
    _sec("Distribuicao de Classificacoes por Indicador")

    _cats  = ["EXCELENTE", "OTIMO", "PARABENS", "ATENCAO", "PRECISA MELHORAR", "S/D"]
    _cores = ["#1e3a5f",   "#2d5a8e", "#4a7db5", "#6b9bd2", "#a0c4ff",           "#94a3b8"]
    _inds  = ["prod_cls",  "efic_cls","rep_cls", "inf_cls"]
    _lbls  = ["Produtividade","Eficacia","Repetida","Infancia"]

    fig_dist = go.Figure()
    for cat, cor in zip(_cats, _cores):
        vals = [
            (df_q[ind].apply(lambda x: x[0]) == cat).sum()
            for ind in _inds
        ]
        if sum(vals) == 0:
            continue
        fig_dist.add_trace(go.Bar(
            name=cat, x=_lbls, y=vals,
            marker_color=cor,
            text=[v if v > 0 else "" for v in vals],
            textposition="auto", textfont_size=11,
        ))

    layout_dist = _lyt("Distribuicao por Indicador e Classificacao", 340)
    layout_dist["barmode"] = "stack"
    layout_dist["yaxis"] = dict(title="Qtd. Tecnicos", gridcolor=C["grid"])
    layout_dist["legend"] = dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5)
    fig_dist.update_layout(layout_dist)
    st.plotly_chart(fig_dist, use_container_width=True)

    _sec("Gerar Atas de Qualidade — Assinatura Digital")
    st.caption("Clique em 📄 ATA para abrir o formulário de assinatura logo abaixo. Ao salvar, o PDF será enviado automaticamente para o Google Drive.")

    _h1, _h2, _h3, _h4, _h5, _h6, _h7 = st.columns([3, 1, 1, 1, 1, 1, 1])
    for col, txt in zip(
        [_h1, _h2, _h3, _h4, _h5, _h6, _h7],
        ["**Tecnico**", "**Prod.**", "**Efic.**", "**Repet.**", "**Inf.**", "**Nota**", ""],
    ):
        col.markdown(txt)

    sup_nome = f.get("supervisor", "N/I") or "N/I"

    if "ata_html" not in st.session_state:
        st.session_state.ata_html = None
        st.session_state.ata_nome = ""

    for _, r in df_q.iterrows():
        pv = f"{r['prod_media']:.1f} /dia" if r['prod_media'] is not None else "S/D"
        ev = f"{r['efic_pct']:.1f}%"       if r['efic_pct']   is not None else "S/D"
        rv = f"{r['rep_pct']:.1f}%"        if r['rep_pct']    is not None else "S/D"
        iv = f"{r['inf_pct']:.1f}%"        if r['inf_pct']    is not None else "S/D"

        c1, c2, c3, c4, c5, c6, c7 = st.columns([3, 1, 1, 1, 1, 1, 1])

        with c1:
            st.write(f"**{r['nome']}** `{r['cod']}`")
        with c2:
            pc = r["prod_cls"]
            st.markdown(f'<span style="color:{pc[2]};font-weight:700">{pc[1]} {pv}</span>',
                        unsafe_allow_html=True)
        with c3:
            ec = r["efic_cls"]
            st.markdown(f'<span style="color:{ec[2]};font-weight:700">{ec[1]} {ev}</span>',
                        unsafe_allow_html=True)
        with c4:
            rc = r["rep_cls"]
            st.markdown(f'<span style="color:{rc[2]};font-weight:700">{rc[1]} {rv}</span>',
                        unsafe_allow_html=True)
        with c5:
            ic = r["inf_cls"]
            st.markdown(f'<span style="color:{ic[2]};font-weight:700">{ic[1]} {iv}</span>',
                        unsafe_allow_html=True)
        with c6:
            cor_n = _cor_nota(r["nota"])
            st.markdown(
                f'<span style="color:{cor_n};font-weight:800;font-size:15px">{r["nota"]}/16</span>',
                unsafe_allow_html=True)
        with c7:
            if st.button("📄 ATA", key=f"ata_q_{r['cod']}"):
                st.session_state.ata_html = _html_ata_qualidade(
                    nome      = r["nome"],
                    codigo    = r["cod"],
                    supervisor= sup_nome,
                    mes_ref   = mes_ref,
                    prod_val  = pv,
                    efic_val  = ev,
                    rep_val   = rv,
                    inf_val   = iv,
                    prod_cls  = r["prod_cls"],
                    efic_cls  = r["efic_cls"],
                    rep_cls   = r["rep_cls"],
                    inf_cls   = r["inf_cls"],
                    nota      = r["nota"],
                )
                st.session_state.ata_nome = r["nome"]
                st.rerun()

    if st.session_state.ata_html:
        st.markdown(f"**📝 Assinatura de {st.session_state.ata_nome}**")
        st.components.v1.html(st.session_state.ata_html, height=900, scrolling=True)
        if st.button("❌ Fechar ATA"):
            st.session_state.ata_html = None
            st.session_state.ata_nome = ""
            st.rerun()


# =============================================================================
# 9. TELA JUSTIFICATIVAS
# =============================================================================

def tela_justificativas(df, f):
    _header("📝", "Justificativas Diárias", f)
    ds = _escopo(df, f)

    datas_disp = sorted(ds["FIM_DT"].dropna().dt.date.unique(), reverse=True)
    if not datas_disp:
        st.warning("Nenhuma data disponível.")
        return

    hoje = datetime.today().date()
    valor_padrao = hoje if hoje in datas_disp else datas_disp[0]
    dia_sel = st.date_input("Data de referência", value=valor_padrao,
                            min_value=datas_disp[-1], max_value=datas_disp[0],
                            key="dia_just")

    dm = ds[ds["FIM_DT"].dt.date == dia_sel].copy()
    data_ref = pd.Timestamp(dia_sel)

    def _endereco(row):
        return _endereco_completo(row)

    def _hist_gpon_just(gpon):
        """Histórico numerado dos últimos 30 dias."""
        cut = data_ref - pd.DateOffset(days=30)
        hist = df[(df["FSLOI_GPONAccess"] == gpon) & (df["FIM_DT"] >= cut) & (df["FIM_DT"] <= data_ref)].sort_values("FIM_DT")
        if hist.empty:
            return "Sem reparos nos últimos 30 dias."
        linhas = []
        for i, (_, row) in enumerate(hist.iterrows(), 1):
            sa = row.get("Número SA", "")
            data = row.get("FIM_DT", "")
            data_str = pd.to_datetime(data).strftime("%d/%m/%Y") if pd.notna(data) else "S/D"
            tecnico = row.get("NOME_TEC", "")
            tr = row.get("CODIGO_TECNICO_EXTRAIDO", "")
            endereco = _endereco(row)
            desc = row.get("Descrição", "")
            linha = f"{i}º: SA {sa} - {data_str} - {tecnico} - {tr}"
            if endereco != "Endereço não disponível":
                linha += f"\n     🏠 {endereco}"
            if desc and str(desc).strip():
                linha += f"\n     📝 {str(desc).strip()}"
            linhas.append(linha)
        return "\n".join(linhas)

    tab1, tab2, tab3, tab4 = st.tabs(["⚠️ PZERO", "🔁 REPETIDOS", "👶 INFÂNCIA", "🚨 ALARMES 24H"])

    supervisor_atual = f.get("supervisor", "N/I")

    for tab, tipo, mask_condition in [
        (tab1, "PZERO", dm[((dm["FLAG_P0_10_DIA"] == "SIM") | (dm["FLAG_P0_15_DIA"] == "SIM"))]),
        (tab2, "REPETIDO", dm[(dm["Macro Atividade"] == "REP-FTTH") & (dm["FLAG_REPETIDO"] == "SIM") & (dm["TEC_ANTERIOR"].notna())]),
        (tab3, "INFANCIA", dm[(dm["Macro Atividade"] == "INST-FTTH") & (dm["FLAG_INFANCIA"] == "SIM")]),
        (tab4, "ALARME_24H", dm[(dm["ALARMADO"] == "SIM")]),
    ]:
        with tab:
            st.markdown(f"### {tipo.replace('_',' ').title()} do dia")
            if mask_condition.empty:
                st.success(f"Nenhum {tipo.lower()} registrado.")
                continue
            for idx, row in mask_condition.iterrows():
                with st.form(key=f"just_{tipo}_{idx}"):
                    c1, c2, c3 = st.columns([2,2,2])
                    c1.write(f"**SA:** {row['Número SA']}")
                    c2.write(f"**TR:** {row['CODIGO_TECNICO_EXTRAIDO']} - {row['NOME_TEC']}")
                    c3.write(f"**GPON:** {row['FSLOI_GPONAccess']}")
                    endereco = _endereco(row)
                    st.caption(f"🏠 Endereço: {endereco}")
                    desc = row.get("Descrição", "")
                    if pd.notna(desc) and str(desc).strip():
                        st.caption(f"📋 Descrição: {str(desc).strip()}")
                    with st.expander("📜 Histórico do GPON (30 dias)"):
                        st.markdown(_hist_gpon_just(row["FSLOI_GPONAccess"]), unsafe_allow_html=True)
                    just = st.text_area("Justificativa", key=f"just_txt_{tipo}_{idx}")
                    if st.form_submit_button("Enviar"):
                        if not just:
                            st.error("Justificativa é obrigatória.")
                        else:
                            payload = {
                                "supervisor": supervisor_atual,
                                "tecnico": row["CODIGO_TECNICO_EXTRAIDO"],
                                "tipo": tipo,
                                "referencia": str(row["Número SA"]),
                                "justificativa": f"[{dia_sel.strftime('%d/%m/%Y')}] {just}"
                            }
                            with st.spinner("Enviando..."):
                                try:
                                    resp = requests.post(URL_JUSTIFICATIVA, json=payload, timeout=15)
                                    if resp.status_code == 200 and resp.json().get("status") == "ok":
                                        st.success(f"✅ {tipo} justificado!")
                                        st.balloons()
                                    else:
                                        st.error(f"❌ Erro: {resp.json().get('mensagem', 'Falha desconhecida')}")
                                except Exception as e:
                                    st.error(f"❌ Falha na comunicação: {e}")
=======
        st.error("Mês inválido.")
        return
    hoje = datetime.now()
    dias_mes = _cal.monthrange(ano, mes)[1]
    dia_max = hoje.day if (ano == hoje.year and mes == hoje.month) else dias_mes
    dias_range = list(range(1, dia_max + 1))
    meses_pt = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    lbl_mes = f"{meses_pt[mes-1]}/{ano}"
    df_valido = df.dropna(subset=["FIM_DT"])
    df_mes = df_valido[(df_valido["FIM_DT"].dt.year == ano) & (df_valido["FIM_DT"].dt.month == mes)].copy()
    df_mes["DIA"] = df_mes["FIM_DT"].dt.day.astype(int)
    suc_m = df_mes[df_mes["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO"]
    semsuc_m = df_mes[df_mes["ESTADO_NORM"] == "CONCLUÍDO SEM SUCESSO"]
    if suc_m.empty:
        st.warning(f"Nenhuma atividade concluída com sucesso em {lbl_mes}.")
        return
    tecs = suc_m["COD_TEC"].nunique()
    total_concluidas = len(suc_m)
    inst_count = (suc_m["MACRO_NORM"] == "INST-FTTH").sum()
    rep_count = (suc_m["MACRO_NORM"] == "REP-FTTH").sum()
    outros_count = total_concluidas - inst_count - rep_count
    cols = st.columns(5)
    for col, (lb, vl, sb, cl) in zip(cols, [("Técnicos Ativos", f"{tecs}", lbl_mes, "kpi-blue"), ("Total Concluídos", f"{total_concluidas:,}", "c/ sucesso", "kpi-blue"), ("INST", f"{inst_count:,}", "Instalações", "kpi-blue"), ("REP", f"{rep_count:,}", "Reparos", "kpi-purple"), ("Outros", f"{outros_count:,}", "MUD/Outros", "kpi-yellow")]):
        col.markdown(_kpi(lb, vl, sb, cl), unsafe_allow_html=True)
    st.write("")
    suc_range = suc_m[suc_m["DIA"].isin(dias_range)]
    if suc_range.empty:
        st.warning("Nenhuma atividade no período selecionado.")
        return
    pivot = suc_range.groupby(["COD_TEC", "DIA"]).size().unstack(fill_value=0)
    for d in dias_range:
        if d not in pivot.columns:
            pivot[d] = 0
    pivot = pivot[dias_range]
    nomes = suc_m.groupby("COD_TEC")["NOME_TEC"].first().fillna("").apply(lambda x: x.title() if x else "")
    ss_tec = semsuc_m.groupby("COD_TEC").size().reindex(pivot.index, fill_value=0)
    dias_trab = suc_range.groupby("COD_TEC")["DIA"].nunique().reindex(pivot.index, fill_value=0)
    pivot.insert(0, "Nome", nomes)
    pivot["Total"] = pivot[dias_range].sum(axis=1)
    pivot["Sem Suc."] = ss_tec.values
    pivot["Dias"] = dias_trab.values
    pivot["Media"] = (pivot["Total"] / pivot["Dias"].replace(0, 1)).round(1)
    pivot["Eficácia%"] = (pivot["Total"] / (pivot["Total"] + pivot["Sem Suc."]).replace(0, 1) * 100).round(1)
    pivot = pivot.sort_values("Total", ascending=False).reset_index().rename(columns={"COD_TEC": "TR"})
    cols_order = ["Nome", "TR"] + dias_range + ["Dias", "Total", "Sem Suc.", "Media", "Eficácia%"]
    pivot = pivot[cols_order]
    def _cor_cel(v):
        try:
            v = int(v)
        except: return ""
        if v >= 8: return "background-color:#7c3aed;color:white;font-weight:700;text-align:center"
        if v in (6,7): return "background-color:#000000;color:white;font-weight:700;text-align:center"
        if v == 5: return "background-color:#1e40af;color:white;font-weight:700;text-align:center"
        if v == 4: return "background-color:#d4edda;color:#155724;font-weight:600;text-align:center"
        if v == 3: return "background-color:#fff3cd;color:#856404;font-weight:600;text-align:center"
        if v in (1,2): return "background-color:#ffcccc;color:#721c24;text-align:center"
        return "background-color:#f0f0f0;color:#adb5bd;text-align:center"
    styled = pivot.style.map(_cor_cel, subset=dias_range)
    altura_total = 38 + (len(pivot) * 35)
    st.dataframe(styled, use_container_width=True, hide_index=True, height=altura_total)

def tela_producao(dm, ds, f):
    _header("📊", "Produção Diária", f)
    suc = dm[dm["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO"]
    conc = dm[dm["ESTADO_NORM"].isin(["CONCLUÍDO COM SUCESSO", "CONCLUÍDO SEM SUCESSO"])]
    inst = suc[suc["MACRO_NORM"] == "INST-FTTH"]
    rep = suc[suc["MACRO_NORM"] == "REP-FTTH"]
    dias = suc["DIA_FIM"].nunique()
    efic = round(len(suc) / len(conc) * 100, 1) if len(conc) > 0 else 0
    med = round(len(suc) / dias, 1) if dias > 0 else 0
    cols = st.columns(6)
    for col, (lb, vl, sb, cl) in zip(cols, [("Concluídos", f"{len(suc):,}", "c/ sucesso", "kpi-blue"), ("Eficácia", f"{efic}%", "suc/total", "kpi-green" if efic>=85 else "kpi-yellow" if efic>=70 else "kpi-red"), ("Instalações", f"{len(inst):,}", "INST-FTTH", "kpi-blue"), ("Reparos", f"{len(rep):,}", "REP-FTTH", "kpi-purple"), ("Dias Trab.", f"{dias}", "c/ encerramento", "kpi-blue"), ("Média/Dia", f"{med}", "atividades/dia", "kpi-blue")]):
        col.markdown(_kpi(lb, vl, sb, cl), unsafe_allow_html=True)
    st.write("")
    _sec("Produção por Dia")
    if not suc.empty:
        prod = suc.groupby("DIA_FIM").agg(Total=("SA", "count"), Inst=("MACRO_NORM", lambda x: (x == "INST-FTTH").sum()), Rep=("MACRO_NORM", lambda x: (x == "REP-FTTH").sum())).reset_index()
        sem_suc = dm[dm["ESTADO_NORM"] == "CONCLUÍDO SEM SUCESSO"]
        if not sem_suc.empty:
            ss = sem_suc.groupby("DIA_FIM").size().reset_index(name="SS")
            prod = prod.merge(ss, on="DIA_FIM", how="left")
        else:
            prod["SS"] = 0
        prod["SS"] = prod["SS"].fillna(0).astype(int)
        prod["Efic%"] = (prod["Total"] / (prod["Total"] + prod["SS"]) * 100).round(1)
        prod["Dia"] = pd.to_datetime(prod["DIA_FIM"]).dt.strftime("%d/%m")
        fig = go.Figure()
        fig.add_bar(x=prod["Dia"], y=prod["Inst"], name="INST", marker_color=ROYAL)
        fig.add_bar(x=prod["Dia"], y=prod["Rep"], name="REP", marker_color=ROYAL_LIGHT)
        fig.update_layout(barmode="stack", showlegend=True, **_lyt("Produção Diária - INST + REP", 320))
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(prod[["Dia","Total","Inst","Rep","SS","Efic%"]].rename(columns={"Inst":"INST","Rep":"REP","SS":"Sem Suc.","Efic%":"Eficácia%"}), use_container_width=True, hide_index=True)
    _sec("Pareto de Técnicos")
    pt = suc.groupby(["COD_TEC","NOME_TEC"]).size().reset_index(name="Prod")
    pt = pt.sort_values("Prod", ascending=False).reset_index(drop=True)
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Top 5 — Maior Produção**")
        t5 = pt.head(5)
        if not t5.empty:
            st.plotly_chart(_bar_h(t5["NOME_TEC"], t5["Prod"], ROYAL, labels=t5["COD_TEC"], h=260), use_container_width=True)
    with c2:
        st.markdown("**Top 5 — Menor Produção**")
        b5 = pt.tail(5).sort_values("Prod")
        if not b5.empty:
            st.plotly_chart(_bar_h(b5["NOME_TEC"], b5["Prod"], ROYAL_LIGHT, labels=b5["COD_TEC"], h=260), use_container_width=True)

def tela_repetidos_vip(df_base, df_vip_rep, f):
    """Tela de repetidos usando a base VIP (oficial)."""
    _header("🔁", "Repetidos (VIP Oficial)", f)
    if df_vip_rep is None:
        st.error("❌ Base VIP de repetidos não disponível para o mês selecionado.")
        return
    # Filtrar pelos técnicos da equipe (se houver filtro)
    tecs_filtro = f["tecs_sup"] if f["tecs_sup"] else []
    if tecs_filtro:
        mask_vip = df_vip_rep['tecnico_anterior'].astype(str).str.upper().isin([t.upper() for t in tecs_filtro])
    else:
        mask_vip = pd.Series(True, index=df_vip_rep.index)
    df_rep = df_vip_rep[mask_vip].copy()
    if df_rep.empty:
        st.info("Nenhum repetido VIP para os filtros selecionados.")
        return
    # Estatísticas
    total_repetidos = len(df_rep)
    # Total de reparos da equipe (BASEBOT) no mês
    df_base_mes = df_base[df_base["MES_REF"] == f["mes"]].copy()
    if tecs_filtro:
        mask_base = df_base_mes['COD_TEC'].isin(tecs_filtro)
    else:
        mask_base = pd.Series(True, index=df_base_mes.index)
    reparos_equipe = df_base_mes[mask_base & (df_base_mes["MACRO_NORM"] == "REP-FTTH") & (df_base_mes["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO")]
    total_reparos = len(reparos_equipe)
    taxa = round(total_repetidos / total_reparos * 100, 2) if total_reparos > 0 else 0
    cols = st.columns(3)
    cols[0].markdown(_kpi("Repetidos (VIP)", f"{total_repetidos}", "GPONs únicos", "kpi-red" if taxa>9 else "kpi-yellow"), unsafe_allow_html=True)
    cols[1].markdown(_kpi("Total Reparos (BASEBOT)", f"{total_reparos}", "concluídos no mês", "kpi-blue"), unsafe_allow_html=True)
    cols[2].markdown(_kpi("Taxa %", f"{taxa}%", "meta: ≤ 9%", "kpi-red" if taxa>9 else "kpi-green"), unsafe_allow_html=True)
    st.write("")
    # Tabela por técnico (tecnico_anterior)
    _sec("Repetidos por Técnico (como PAI)")
    rep_por_tec = df_rep.groupby('tecnico_anterior').size().reset_index(name='Repetidos')
    # Adicionar nome (buscar no BASEBOT)
    nomes_tec = df_base.drop_duplicates(subset=['COD_TEC']).set_index('COD_TEC')['NOME_TEC'].to_dict()
    rep_por_tec['Nome'] = rep_por_tec['tecnico_anterior'].map(nomes_tec).fillna(rep_por_tec['tecnico_anterior'])
    rep_por_tec = rep_por_tec.rename(columns={'tecnico_anterior': 'TR'})
    rep_por_tec = rep_por_tec[['Nome', 'TR', 'Repetidos']].sort_values('Repetidos', ascending=False)
    st.dataframe(rep_por_tec, use_container_width=True, hide_index=True)
    # Detalhamento dos GPONs repetidos
    _sec("GPONs com repetição")
    # Agrupar por GPON e mostrar quantas repetições e técnicos envolvidos
    gpons = df_rep.groupby('gpon').agg(Repeticoes=('gpon', 'count'), Tecnicos=list).reset_index()
    st.dataframe(gpons, use_container_width=True, hide_index=True)

def tela_infancia_vip(df_base, df_vip_inf, f):
    """Tela de infância usando a base VIP oficial."""
    _header("👶", "Infância (VIP Oficial)", f)
    if df_vip_inf is None:
        st.error("❌ Base VIP de infância não disponível para o mês selecionado.")
        return
    tecs_filtro = f["tecs_sup"] if f["tecs_sup"] else []
    if tecs_filtro:
        mask_vip = df_vip_inf['tecnico_anterior'].astype(str).str.upper().isin([t.upper() for t in tecs_filtro])
    else:
        mask_vip = pd.Series(True, index=df_vip_inf.index)
    df_inf = df_vip_inf[mask_vip].copy()
    if df_inf.empty:
        st.info("Nenhuma infância VIP para os filtros selecionados.")
        return
    total_inf = len(df_inf)
    # Total de instalações da equipe (BASEBOT)
    df_base_mes = df_base[df_base["MES_REF"] == f["mes"]].copy()
    if tecs_filtro:
        mask_base = df_base_mes['COD_TEC'].isin(tecs_filtro)
    else:
        mask_base = pd.Series(True, index=df_base_mes.index)
    instalacoes = df_base_mes[mask_base & (df_base_mes["MACRO_NORM"] == "INST-FTTH") & (df_base_mes["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO")]
    total_inst = len(instalacoes)
    taxa = round(total_inf / total_inst * 100, 2) if total_inst > 0 else 0
    cols = st.columns(3)
    cols[0].markdown(_kpi("Infância (VIP)", f"{total_inf}", "GPONs únicos", "kpi-red" if taxa>5 else "kpi-yellow"), unsafe_allow_html=True)
    cols[1].markdown(_kpi("Total Instalações", f"{total_inst}", "concluídas no mês", "kpi-blue"), unsafe_allow_html=True)
    cols[2].markdown(_kpi("Taxa %", f"{taxa}%", "meta: ≤ 5%", "kpi-red" if taxa>5 else "kpi-green"), unsafe_allow_html=True)
    st.write("")
    _sec("Infância por Técnico (como instalador)")
    inf_por_tec = df_inf.groupby('tecnico_anterior').size().reset_index(name='Infância')
    nomes_tec = df_base.drop_duplicates(subset=['COD_TEC']).set_index('COD_TEC')['NOME_TEC'].to_dict()
    inf_por_tec['Nome'] = inf_por_tec['tecnico_anterior'].map(nomes_tec).fillna(inf_por_tec['tecnico_anterior'])
    inf_por_tec = inf_por_tec.rename(columns={'tecnico_anterior': 'TR'})
    inf_por_tec = inf_por_tec[['Nome', 'TR', 'Infância']].sort_values('Infância', ascending=False)
    st.dataframe(inf_por_tec, use_container_width=True, hide_index=True)
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)

def tela_qualidade(dm, ds, f):
    _header("🏆", "Qualidade", f)
    tecs = dm["COD_TEC"].dropna().unique()
    if len(tecs) == 0:
        st.warning("Nenhum técnico encontrado.")
        return
    rows = []
    for cod in tecs:
        df_tec = dm[dm["COD_TEC"] == cod].copy()
        if df_tec.empty: continue
        nome = df_tec["NOME_TEC"].iloc[0] if not df_tec.empty else cod
        suc = df_tec[df_tec["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO"]
        dias_unicos = suc["DIA_FIM"].nunique()
        prod_media = round(len(suc) / dias_unicos, 1) if dias_unicos > 0 else None
        n_suc = len(suc)
        n_ss = len(df_tec[df_tec["ESTADO_NORM"] == "CONCLUÍDO SEM SUCESSO"])
        efic = round(n_suc / (n_suc + n_ss) * 100, 1) if (n_suc + n_ss) > 0 else None
        rep_den = len(df_tec[df_tec["MACRO_NORM"] == "REP-FTTH"])
        rep_num = len(df_tec[(df_tec["MACRO_NORM"] == "REP-FTTH") & (df_tec["FLAG_REPETIDO"] == "SIM")])
        rep_pct = round(rep_num / rep_den * 100, 1) if rep_den > 0 else None
        inst_den = len(df_tec[(df_tec["MACRO_NORM"] == "INST-FTTH") & (df_tec["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO")])
        inst_num = len(df_tec[(df_tec["MACRO_NORM"] == "INST-FTTH") & (df_tec["ESTADO_NORM"] == "CONCLUÍDO COM SUCESSO") & (df_tec["FLAG_INFANCIA"] == "SIM")])
        inf_pct = round(inst_num / inst_den * 100, 1) if inst_den > 0 else None
        prod_cls = ("S/D","⬜","#94a3b8",0) if prod_media is None else (("EXCELENTE","🏆","#1e3a5f",4) if prod_media >= 6 else ("PARABENS","🟢","#16a34a",3) if prod_media >= 4 else ("ATENCAO","🟡","#d97706",2) if prod_media >= 3 else ("PRECISA MELHORAR","🔴","#dc2626",1))
        efic_cls = ("S/D","⬜","#94a3b8",0) if efic is None else (("EXCELENTE","🏆","#1e3a5f",4) if efic >= 85 else ("PARABENS","🟢","#16a34a",3) if efic >= 82 else ("ATENCAO","🟡","#d97706",2) if efic >= 75 else ("PRECISA MELHORAR","🔴","#dc2626",1))
        rep_cls = ("S/D","⬜","#94a3b8",0) if rep_pct is None else (("EXCELENTE","🏆","#1e3a5f",4) if rep_pct < 2 else ("OTIMO","🟢","#16a34a",3) if rep_pct <= 5 else ("PARABENS","🟢","#16a34a",2) if rep_pct <= 9 else ("PRECISA MELHORAR","🔴","#dc2626",0))
        inf_cls = ("S/D","⬜","#94a3b8",0) if inf_pct is None else (("OTIMO","🟢","#16a34a",3) if inf_pct < 3 else ("PRECISA MELHORAR","🔴","#dc2626",0))
        nota = prod_cls[3] + efic_cls[3] + rep_cls[3] + inf_cls[3]
        rows.append({"cod":cod, "nome":nome, "prod_media":prod_media, "efic_pct":efic, "rep_pct":rep_pct, "inf_pct":inf_pct, "prod_cls":prod_cls, "efic_cls":efic_cls, "rep_cls":rep_cls, "inf_cls":inf_cls, "nota":nota})
    if not rows:
        st.warning("Sem dados suficientes.")
        return
    df_q = pd.DataFrame(rows).sort_values("nota", ascending=False).reset_index(drop=True)
    n_exc = (df_q["nota"] >= 13).sum()
    n_bom = ((df_q["nota"] >= 9) & (df_q["nota"] < 13)).sum()
    n_atc = ((df_q["nota"] >= 5) & (df_q["nota"] < 9)).sum()
    n_mel = (df_q["nota"] < 5).sum()
    cols_kpi = st.columns(5)
    for col, lbl, val, cls in zip(cols_kpi, ["Técnicos","🏆 Excelente","✅ Bom","⚠️ Atenção","🔴 Melhoria"], [len(df_q), n_exc, n_bom, n_atc, n_mel], ["kpi-blue","kpi-blue","kpi-green","kpi-yellow","kpi-red"]):
        col.markdown(_kpi(lbl, val, "", cls), unsafe_allow_html=True)
    _sec("Ranking de Qualidade")
    _cor_map = {"EXCELENTE":("🏆","#1e3a5f"), "PARABENS":("🟢","#16a34a"), "OTIMO":("🟢","#15803d"), "ATENCAO":("🟡","#d97706"), "PRECISA MELHORAR":("🔴","#dc2626"), "S/D":("⬜","#94a3b8")}
    html_table = '<table class="qualidade-table"><thead><tr><th>Nome</th><th>TR</th><th style="text-align:center">Prod.</th><th style="text-align:center">Efic.</th><th style="text-align:center">Repet.</th><th style="text-align:center">Infan.</th><th style="text-align:center">Nota</th></tr></thead><tbody>'
    for _, r in df_q.iterrows():
        pv = f"{r['prod_media']:.1f}" if r['prod_media'] is not None else "S/D"
        ev = f"{r['efic_pct']:.1f}%" if r['efic_pct'] is not None else "S/D"
        rv = f"{r['rep_pct']:.1f}%" if r['rep_pct'] is not None else "S/D"
        iv = f"{r['inf_pct']:.1f}%" if r['inf_pct'] is not None else "S/D"
        nota_num = r['nota']
        nota_cor = "#1e3a5f" if nota_num >= 13 else "#16a34a" if nota_num >= 9 else "#d97706" if nota_num >= 5 else "#dc2626"
        prod_emoji, prod_color = _cor_map.get(r["prod_cls"][0], ("","#94a3b8"))
        efic_emoji, efic_color = _cor_map.get(r["efic_cls"][0], ("","#94a3b8"))
        rep_emoji, rep_color = _cor_map.get(r["rep_cls"][0], ("","#94a3b8"))
        inf_emoji, inf_color = _cor_map.get(r["inf_cls"][0], ("","#94a3b8"))
        html_table += f'<tr><td><strong>{r["nome"]}</strong></td><td style="color:#64748b;font-family:monospace">{r["cod"]}</td><td style="text-align:center;color:{prod_color};font-weight:600">{prod_emoji} {pv}</td><td style="text-align:center;color:{efic_color};font-weight:600">{efic_emoji} {ev}</td><td style="text-align:center;color:{rep_color};font-weight:600">{rep_emoji} {rv}</td><td style="text-align:center;color:{inf_color};font-weight:600">{inf_emoji} {iv}</td><td style="text-align:center;font-weight:700;color:{nota_cor}">{r["nota"]}/16</td></tr>'
    html_table += '</tbody></table>'
    st.markdown(html_table, unsafe_allow_html=True)

# =============================================================================
# 8. MAIN
# =============================================================================

def main():
<<<<<<< HEAD
    df = carregar_base()
    if df is None:
        st.error("❌ Não foi possível carregar a base de dados. Verifique o token de acesso.")
        return

    f    = sidebar(df)
    tela = f["tela"]
    dm   = _filtrar(df, f)
    ds   = _escopo(df, f)

    if f["supervisor"]:
        st.markdown(
            f'<div class="banner-sup">👑 Equipe de <strong>{f["supervisor"]}</strong>'
            f' — {len(f["tecs_sup"])} tecnico(s) | {f["mes"]}</div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            '<div class="banner-sup" style="background:#f0f4f8">'
            '🕐 Dados carregados com sucesso.</div>',
            unsafe_allow_html=True)

    if tela == "📝 Justificativas":
        tela_justificativas(df, f)
        return

    if dm.empty and tela not in ("📅 Diario", "📆 Calendario"):
        st.warning("Nenhum dado encontrado para os filtros selecionados.")
        return

    if tela == "📅 Diario":
        tela_diario(df, ds, f)
    elif tela == "📊 Producao Diaria":
        tela_producao(dm, ds, f)
    elif tela == "🔁 Repetidos":
        tela_repetidos(dm, ds, f)
    elif tela == "👶 Infancia":
        tela_infancia(dm, ds, f)
    elif tela == "📆 Calendario":
        tela_calendario(df, ds, f)
    elif tela == "🏆 Qualidade":
        tela_qualidade(dm, ds, f)

=======
    try:
        # Carrega as bases (BASEBOT + VIPs)
        df_base, df_vip_rep, df_vip_inf = carregar_todas_bases(datetime.now().strftime("%Y-%m"))
        if df_base is None:
            st.error("❌ Falha ao carregar BASEBOT.csv")
            return
        df_base, ultima_data = processar_basebot(df_base)
        f = sidebar(df_base, ultima_data)
        if not f:
            return
        dm = _filtrar(df_base, f)
        ds = _escopo(df_base, f)
        if f["supervisor"]:
            st.markdown(f'<div class="info-box">👑 Equipe de <strong>{f["supervisor"]}</strong> — {len(f["tecs_sup"])} técnico(s) | {f["mes"]}</div>', unsafe_allow_html=True)
        tela = f["tela"]
        if tela == "📅 Diario":
            tela_diario(ds, f)
        elif tela == "📊 Producao Diaria":
            tela_producao(dm, ds, f)
        elif tela == "🔁 Repetidos":
            tela_repetidos_vip(df_base, df_vip_rep, f)
        elif tela == "👶 Infancia":
            tela_infancia_vip(df_base, df_vip_inf, f)
        elif tela == "📆 Calendario":
            tela_calendario(ds, f)
        elif tela == "🏆 Qualidade":
            tela_qualidade(dm, ds, f)
    except Exception as e:
        st.error("❌ Ocorreu um erro")
        st.code(traceback.format_exc())
>>>>>>> e324f7a (Painel com MES_REFERENCIA, sem cache e endereço completo)

if __name__ == "__main__":
    main()
